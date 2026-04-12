from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.db import transaction
from django.db.models import Sum, Count, Q, F, Prefetch
from django.utils import timezone
from django.urls import reverse
from django.views.decorators.csrf import csrf_protect,csrf_exempt
from django.views.decorators.http import require_http_methods, require_POST, require_GET
from django.core.paginator import Paginator
from django.template.loader import render_to_string
from django.conf import settings
from decimal import Decimal
import json
import logging
from datetime import datetime, timedelta

# Models
from .models import (
    # Cadastros
    Cliente, Fornecedor, Produto, Vendedor, Empresa,  
    CategoriaProduto, UnidadeMedida, Projeto,
    Categoria, Transportadora, Marca,  # ✅ ADICIONADOS: Categoria, Transportadora
    
    # Compras
    CotacaoMae, ItemSolicitado, PedidoCompra, ItemPedidoCompra, 
    NotaFiscalEntrada, ItemNotaFiscalEntrada, RegraAprovacao, PedidoAprovacao,
    
    # Vendas  
    Orcamento, ItemOrcamento, OrcamentoProjeto, PedidoVenda, ItemPedidoVenda,
    NotaFiscalSaida, ItemNotaFiscalSaida,
    
    # Financeiro
    ContaPagar, ContaReceber, MovimentoCaixa, ConfiguracaoDRE, 
    ItemRelatorioDRE, ExtratoBancario, 
    CategoriaFinanceira, CentroCusto, OrcamentoFinanceiro,
    LancamentoExtrato, ContaBancaria,
    FluxoCaixa, ConciliacaoBancaria, PlanejadoRealizado,  # ✅ ADICIONADOS
    
    # Estoque
    MovimentacaoEstoque, ItemMovimentacaoEstoque, Inventario, ItemInventario,
    TransferenciaEstoque, ItemTransferencia, Deposito, SaldoEstoque,
    EntradaNFE, ItemEntradaNFE, PosicaoEstoque,
    
    # Pagamentos
    CondicaoPagamento, ItemCondicaoPagamento, FormaPagamento,
)

# Forms
from .forms import (
    ClienteForm, FornecedorForm, ProdutoForm, VendedorForm, EmpresaForm,
    TransportadoraForm, CategoriaForm, MarcaForm, UnidadeMedidaForm, ProjetoForm,
    CotacaoForm, ItemCotacaoForm, PedidoCompraForm, ItemPedidoCompraForm,
    NotaFiscalEntradaForm, ItemNotaFiscalEntradaForm,
    OrcamentoForm, ItemOrcamentoForm, OrcamentoProjetoForm,
    PedidoVendaForm, ItemPedidoVendaForm,
    NotaFiscalSaidaForm, ItemNotaFiscalSaidaForm,
    ContaPagarForm, ContaReceberForm, FluxoCaixaForm,
    LancamentoDREForm, ConciliacaoBancariaForm, PlanejadoRealizadoForm,
    CategoriaFinanceiraForm, CentroCustoForm, ContaBancariaForm,
    FormaPagamentoForm, CondicaoPagamentoForm,
    MovimentacaoEstoqueForm, ItemMovimentacaoEstoqueForm,
    InventarioForm, ItemInventarioForm,
    TransferenciaEstoqueForm, ItemTransferenciaEstoqueForm,
    DepositoForm, RegraAprovacaoForm, PedidoAprovacaoForm
)

logger = logging.getLogger(__name__)


# =============================================================================
# 1. CADASTROS
# =============================================================================

# -----------------------------------------------------------------------------
# CLIENTES
# -----------------------------------------------------------------------------

@login_required
def cliente_manager(request):
    """Lista de clientes com filtros e paginação"""
    queryset = Cliente.objects.all().order_by('-id')

    # Filtros
    search = request.GET.get('search', '')
    if search:
        queryset = queryset.filter(
            Q(nome__icontains=search) | 
            Q(cnpj_cpf__icontains=search) |
            Q(email__icontains=search)
        )

    status = request.GET.get('status', '')
    if status:
        queryset = queryset.filter(ativo=(status == 'ativo'))

    paginator = Paginator(queryset, 25)
    page = request.GET.get('page', 1)
    clientes = paginator.get_page(page)

    context = {
        'clientes': clientes,
        'search': search,
        'status': status,
        'total': queryset.count()
    }
    return render(request, 'cadastro/cliente_manager.html', context)


@login_required
def cliente_form(request, pk=None):
    """Formulário de cliente (add/edit)"""
    cliente = get_object_or_404(Cliente, pk=pk) if pk else None

    if request.method == 'POST':
        form = ClienteForm(request.POST, instance=cliente)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cliente salvo com sucesso!')
            return redirect('ERP_ServicesBI:cliente_manager')  # ✅ CORRIGIDO
    else:
        form = ClienteForm(instance=cliente)

    context = {
        'form': form,
        'cliente': cliente,
        'is_edit': bool(pk)
    }
    return render(request, 'cadastro/cliente_form.html', context)


@login_required
def cliente_delete(request, pk):
    """Excluir cliente"""
    cliente = get_object_or_404(Cliente, pk=pk)

    if request.method == 'POST':
        cliente.delete()
        messages.success(request, 'Cliente excluído com sucesso!')
        return redirect('ERP_ServicesBI:cliente_manager')

    return render(request, 'cadastro/cliente_confirm_delete.html', {'cliente': cliente})


@login_required
def cliente_detail_api(request, pk):
    """API para detalhes do cliente (usado em selects)"""
    cliente = get_object_or_404(Cliente, pk=pk)
    return JsonResponse({
        'id': cliente.id,
        'nome': cliente.nome,
        'cnpj_cpf': cliente.cnpj_cpf,
        'endereco': cliente.endereco,
        'cidade': cliente.cidade,
        'estado': cliente.estado,
        'cep': cliente.cep,
        'telefone': cliente.telefone,
        'email': cliente.email,
        'condicao_pagamento_padrao_id': cliente.condicao_pagamento_padrao_id,
        'forma_pagamento_padrao_id': cliente.forma_pagamento_padrao_id
    })


# -----------------------------------------------------------------------------
# FORNECEDORES
# -----------------------------------------------------------------------------

@login_required
def fornecedor_manager(request):
    """Lista de fornecedores com paginação e filtros"""
    queryset = Fornecedor.objects.all().order_by('-id')
    
    search = request.GET.get('search', '')
    if search:
        queryset = queryset.filter(
            Q(nome_razao_social__icontains=search) |
            Q(cpf_cnpj__icontains=search)
        )
    
    # Estatísticas para os cards
    total = queryset.count()
    total_ativos = queryset.filter(ativo=True).count()
    total_inativos = queryset.filter(ativo=False).count()
    percentual_ativos = round((total_ativos / total * 100), 1) if total > 0 else 0
    percentual_inativos = round((total_inativos / total * 100), 1) if total > 0 else 0
    
    # Paginação
    paginator = Paginator(queryset, 25)
    page = request.GET.get('page', 1)
    fornecedores = paginator.get_page(page)
    
    return render(request, 'cadastro/fornecedor_manager.html', {
        'fornecedores': fornecedores,
        'search': search,
        'total': total,
        'total_ativos': total_ativos,
        'total_inativos': total_inativos,
        'percentual_ativos': percentual_ativos,
        'percentual_inativos': percentual_inativos,
    })

@login_required
def fornecedor_form(request, pk=None):
    """Formulário de fornecedor"""
    fornecedor = get_object_or_404(Fornecedor, pk=pk) if pk else None

    if request.method == 'POST':
        form = FornecedorForm(request.POST, instance=fornecedor)
        if form.is_valid():
            form.save()
            messages.success(request, 'Fornecedor salvo com sucesso!')
            return redirect('fornecedor_manager')
    else:
        form = FornecedorForm(instance=fornecedor)

    return render(request, 'cadastro/fornecedor_form.html', {
        'form': form,
        'fornecedor': fornecedor,
        'is_edit': bool(pk)
    })


@login_required
def fornecedor_delete(request, pk):
    """Excluir fornecedor"""
    fornecedor = get_object_or_404(Fornecedor, pk=pk)

    if request.method == 'POST':
        fornecedor.delete()
        messages.success(request, 'Fornecedor excluído com sucesso!')
        return redirect('ERP_ServicesBI:fornecedor_manager')

    return render(request, 'cadastro/fornecedor_confirm_delete.html', {'fornecedor': fornecedor})


# -----------------------------------------------------------------------------
# PRODUTOS
# -----------------------------------------------------------------------------

@login_required
def produto_manager(request):
    """Lista de produtos"""
    queryset = Produto.objects.all().order_by('-id')

    search = request.GET.get('search', '')
    if search:
        queryset = queryset.filter(
            Q(codigo__icontains=search) | 
            Q(descricao__icontains=search) |
            Q(sku__icontains=search)
        )

    categoria_id = request.GET.get('categoria', '')
    if categoria_id:
        queryset = queryset.filter(categoria_id=categoria_id)

    paginator = Paginator(queryset, 25)
    page = request.GET.get('page', 1)
    produtos = paginator.get_page(page)

    return render(request, 'cadastro/produto_manager.html', {
        'produtos': produtos,
        'search': search,
        'categorias': Categoria.objects.all(),
        'categoria_id': categoria_id
    })


@login_required
def produto_form(request, pk=None):
    """Formulário de produto"""
    produto = get_object_or_404(Produto, pk=pk) if pk else None

    if request.method == 'POST':
        form = ProdutoForm(request.POST, instance=produto)
        if form.is_valid():
            form.save()
            messages.success(request, 'Produto salvo com sucesso!')
            return redirect('ERP_ServicesBI:produto_manager')
    else:
        form = ProdutoForm(instance=produto)

    return render(request, 'cadastro/produto_form.html', {
        'form': form,
        'produto': produto,
        'is_edit': bool(pk)
    })


@login_required
def produto_delete(request, pk):
    """Excluir produto"""
    produto = get_object_or_404(Produto, pk=pk)

    if request.method == 'POST':
        produto.delete()
        messages.success(request, 'Produto excluído com sucesso!')
        return redirect('ERP_ServicesBI:produto_manager')

    return render(request, 'cadastro/produto_confirm_delete.html', {'produto': produto})


@login_required
def api_produto_saldo_disponivel(request, pk):
    """API para consultar saldo disponível do produto"""
    produto = get_object_or_404(Produto, pk=pk)

    # Calcular saldo
    saldo = PosicaoEstoque.objects.filter(produto=produto).aggregate(
        total=Sum('quantidade')
    )['total'] or 0

    return JsonResponse({
        'id': produto.id,
        'codigo': produto.codigo,
        'descricao': produto.descricao,
        'saldo_disponivel': float(saldo),
        'preco_custo': float(produto.preco_custo) if produto.preco_custo else 0,
        'preco_venda': float(produto.preco_venda) if produto.preco_venda else 0
    })


# -----------------------------------------------------------------------------
# VENDEDORES
# -----------------------------------------------------------------------------

@login_required
def vendedor_manager(request):
    """Lista de vendedores"""
    queryset = Vendedor.objects.all().order_by('-id')
    
    # Filtro de busca
    search = request.GET.get('search', '')
    if search:
        queryset = queryset.filter(
            Q(nome__icontains=search) | 
            Q(email__icontains=search) |
            Q(cpf__icontains=search)
        )
    
    # Estatísticas
    total = queryset.count()
    total_ativos = queryset.filter(ativo=True).count()
    total_inativos = queryset.filter(ativo=False).count()
    
    # Paginação
    paginator = Paginator(queryset, 25)
    page = request.GET.get('page', 1)
    
    try:
        vendedores = paginator.page(page)  # ✅ Isso retorna o objeto Page
    except PageNotAnInteger:
        vendedores = paginator.page(1)
    except EmptyPage:
        vendedores = paginator.page(paginator.num_pages)
    
    return render(request, 'cadastro/vendedor_manager.html', {
        'vendedores': vendedores,  # ✅ Agora tem .paginator.count
        'total': total,
        'total_ativos': total_ativos,
        'total_inativos': total_inativos,
        'search': search,
    })


@login_required
def vendedor_form(request, pk=None):
    """Formulário de vendedor"""
    vendedor = get_object_or_404(Vendedor, pk=pk) if pk else None

    if request.method == 'POST':
        form = VendedorForm(request.POST, instance=vendedor)
        if form.is_valid():
            form.save()
            messages.success(request, 'Vendedor salvo com sucesso!')
            return redirect('ERP_ServicesBI:vendedor_manager')
    else:
        form = VendedorForm(instance=vendedor)

    return render(request, 'cadastro/vendedor_form.html', {
        'form': form,
        'vendedor': vendedor,
        'is_edit': bool(pk)
    })


@login_required
def vendedor_delete(request, pk):
    """Excluir vendedor"""
    vendedor = get_object_or_404(Vendedor, pk=pk)

    if request.method == 'POST':
        vendedor.delete()
        messages.success(request, 'Vendedor excluído com sucesso!')
        return redirect('ERP_ServicesBI:vendedor_manager')

    return render(request, 'cadastro/vendedor_confirm_delete.html', {'vendedor': vendedor})


# -----------------------------------------------------------------------------
# EMPRESAS
# -----------------------------------------------------------------------------

@login_required
def empresa_manager(request):
    """Lista de empresas cadastradas"""
    empresas_list = Empresa.objects.all().order_by('-id')
    
    # Estatísticas para os cards de resumo
    total = empresas_list.count()
    total_ativos = empresas_list.filter(ativo=True).count()
    total_inativos = empresas_list.filter(ativo=False).count()
    
    # Paginação
    per_page = request.GET.get('per_page', 25)
    try:
        per_page = int(per_page)
    except ValueError:
        per_page = 25
    
    paginator = Paginator(empresas_list, per_page)
    page = request.GET.get('page')
    empresas = paginator.get_page(page)
    
    # Cálculos de percentuais
    percentual_ativos = round((total_ativos / total * 100), 1) if total > 0 else 0
    percentual_inativos = round((total_inativos / total * 100), 1) if total > 0 else 0
    
    context = {
        'empresas': empresas,
        'total': total,
        'total_ativos': total_ativos,
        'total_inativos': total_inativos,
        'percentual_ativos': percentual_ativos,
        'percentual_inativos': percentual_inativos,
        'taxa_crescimento': 12,  # Placeholder - pode calcular com base em dados históricos
        'taxa_ativos': 5,        # Placeholder
        'taxa_inativos': 2,      # Placeholder
        'per_page': per_page,
        'search': request.GET.get('search', ''),
    }
    
    return render(request, 'cadastro/empresa_manager.html', context)


@login_required
def empresa_form(request, pk=None):
    """Formulário de empresa"""
    empresa = get_object_or_404(Empresa, pk=pk) if pk else None

    if request.method == 'POST':
        form = EmpresaForm(request.POST, instance=empresa)
        if form.is_valid():
            form.save()
            messages.success(request, 'Empresa salva com sucesso!')
            return redirect('ERP_ServicesBI:empresa_manager')
    else:
        form = EmpresaForm(instance=empresa)

    return render(request, 'cadastro/empresa_form.html', {
        'form': form,
        'empresa': empresa,
        'is_edit': bool(pk)
    })


@login_required
def empresa_delete(request, pk):
    """Excluir empresa"""
    empresa = get_object_or_404(Empresa, pk=pk)

    if request.method == 'POST':
        empresa.delete()
        messages.success(request, 'Empresa excluída com sucesso!')
        return redirect('ERP_ServicesBI:empresa_manager')

    return render(request, 'cadastro/empresa_confirm_delete.html', {'empresa': empresa})


# -----------------------------------------------------------------------------
# TRANSPORTADORAS
# -----------------------------------------------------------------------------

@login_required
def transportadora_manager(request):
    """Lista de transportadoras"""
    transportadoras = Transportadora.objects.all().order_by('-id')
    return render(request, 'cadastro/transportadora_manager.html', {'transportadoras': transportadoras})


@login_required
def transportadora_form(request, pk=None):
    """Formulário de transportadora"""
    transportadora = get_object_or_404(Transportadora, pk=pk) if pk else None

    if request.method == 'POST':
        form = TransportadoraForm(request.POST, instance=transportadora)
        if form.is_valid():
            form.save()
            messages.success(request, 'Transportadora salva com sucesso!')
            return redirect('ERP_ServicesBI:transportadora_manager')
    else:
        form = TransportadoraForm(instance=transportadora)

    return render(request, 'cadastro/transportadora_form.html', {
        'form': form,
        'transportadora': transportadora,
        'is_edit': bool(pk)
    })


@login_required
def transportadora_delete(request, pk):
    """Excluir transportadora"""
    transportadora = get_object_or_404(Transportadora, pk=pk)

    if request.method == 'POST':
        transportadora.delete()
        messages.success(request, 'Transportadora excluída com sucesso!')
        return redirect('ERP_ServicesBI:transportadora_manager')

    return render(request, 'cadastro/transportadora_confirm_delete.html', {'transportadora': transportadora})


# -----------------------------------------------------------------------------
# CATEGORIAS
# -----------------------------------------------------------------------------

@login_required
def categoria_manager(request):
    """Lista de categorias"""
    categorias = Categoria.objects.all().order_by('-id')
    return render(request, 'cadastro/categoria_manager.html', {'categorias': categorias})


@login_required
def categoria_form(request, pk=None):
    """Formulário de categoria"""
    categoria = get_object_or_404(Categoria, pk=pk) if pk else None

    if request.method == 'POST':
        form = CategoriaForm(request.POST, instance=categoria)
        if form.is_valid():
            form.save()
            messages.success(request, 'Categoria salva com sucesso!')
            return redirect('ERP_ServicesBI:categoria_manager')
    else:
        form = CategoriaForm(instance=categoria)

    return render(request, 'cadastro/categoria_form.html', {
        'form': form,
        'categoria': categoria,
        'is_edit': bool(pk)
    })


@login_required
def categoria_delete(request, pk):
    """Excluir categoria"""
    categoria = get_object_or_404(Categoria, pk=pk)

    if request.method == 'POST':
        categoria.delete()
        messages.success(request, 'Categoria excluída com sucesso!')
        return redirect('ERP_ServicesBI:categoria_manager')

    return render(request, 'cadastro/categoria_confirm_delete.html', {'categoria': categoria})


# -----------------------------------------------------------------------------
# MARCAS
# -----------------------------------------------------------------------------

@login_required
def marca_manager(request):
    """Lista de marcas"""
    marcas = Marca.objects.all().order_by('-id')
    return render(request, 'cadastro/marca_manager.html', {'marcas': marcas})


@login_required
def marca_form(request, pk=None):
    """Formulário de marca"""
    marca = get_object_or_404(Marca, pk=pk) if pk else None

    if request.method == 'POST':
        form = MarcaForm(request.POST, instance=marca)
        if form.is_valid():
            form.save()
            messages.success(request, 'Marca salva com sucesso!')
            return redirect('ERP_ServicesBI:marca_manager')
    else:
        form = MarcaForm(instance=marca)

    return render(request, 'cadastro/marca_form.html', {
        'form': form,
        'marca': marca,
        'is_edit': bool(pk)
    })


@login_required
def marca_delete(request, pk):
    """Excluir marca"""
    marca = get_object_or_404(Marca, pk=pk)

    if request.method == 'POST':
        marca.delete()
        messages.success(request, 'Marca excluída com sucesso!')
        return redirect('ERP_ServicesBI:marca_manager')

    return render(request, 'cadastro/marca_confirm_delete.html', {'marca': marca})


# -----------------------------------------------------------------------------
# UNIDADES DE MEDIDA
# -----------------------------------------------------------------------------

@login_required
def unidade_medida_manager(request):
    """Lista de unidades de medida"""
    unidades = UnidadeMedida.objects.all().order_by('-id')
    return render(request, 'cadastro/unidade_medida_manager.html', {'unidades': unidades})


@login_required
def unidade_medida_form(request, pk=None):
    """Formulário de unidade de medida"""
    unidade = get_object_or_404(UnidadeMedida, pk=pk) if pk else None

    if request.method == 'POST':
        form = UnidadeMedidaForm(request.POST, instance=unidade)
        if form.is_valid():
            form.save()
            messages.success(request, 'Unidade de medida salva com sucesso!')
            return redirect('ERP_ServicesBI:unidade_medida_manager')
    else:
        form = UnidadeMedidaForm(instance=unidade)

    return render(request, 'cadastro/unidade_medida_form.html', {
        'form': form,
        'unidade': unidade,
        'is_edit': bool(pk)
    })


@login_required
def unidade_medida_delete(request, pk):
    """Excluir unidade de medida"""
    unidade = get_object_or_404(UnidadeMedida, pk=pk)

    if request.method == 'POST':
        unidade.delete()
        messages.success(request, 'Unidade de medida excluída com sucesso!')
        return redirect('ERP_ServicesBI:unidade_medida_manager')

    return render(request, 'cadastro/unidade_medida_confirm_delete.html', {'unidade': unidade})


# -----------------------------------------------------------------------------
# PROJETOS
# -----------------------------------------------------------------------------

@login_required
def projeto_manager(request):
    """Lista de projetos"""
    projetos = Projeto.objects.all().order_by('-id')
    return render(request, 'cadastro/projeto_manager.html', {'projetos': projetos})


@login_required
def projeto_form(request, pk=None):
    """Formulário de projeto"""
    projeto = get_object_or_404(Projeto, pk=pk) if pk else None

    if request.method == 'POST':
        form = ProjetoForm(request.POST, instance=projeto)
        if form.is_valid():
            form.save()
            messages.success(request, 'Projeto salvo com sucesso!')
            return redirect('ERP_ServicesBI:projeto_manager')
    else:
        form = ProjetoForm(instance=projeto)

    return render(request, 'cadastro/projeto_form.html', {
        'form': form,
        'projeto': projeto,
        'is_edit': bool(pk)
    })


@login_required
def projeto_excluir(request, pk):
    """Excluir projeto"""
    projeto = get_object_or_404(Projeto, pk=pk)

    if request.method == 'POST':
        projeto.delete()
        messages.success(request, 'Projeto excluído com sucesso!')
        return redirect('ERP_ServicesBI:projeto_manager')

    return render(request, 'cadastro/projeto_confirm_delete.html', {'projeto': projeto})


@require_POST
def projeto_create_ajax(request):
    """Criar projeto via AJAX"""
    try:
        data = json.loads(request.body)
        projeto = Projeto.objects.create(
            nome=data.get('nome'),
            descricao=data.get('descricao', ''),
            data_inicio=data.get('data_inicio'),
            data_fim_prevista=data.get('data_fim_prevista'),
            responsavel_id=data.get('responsavel_id'),
            status='ativo'
        )
        return JsonResponse({'success': True, 'id': projeto.id, 'nome': projeto.nome})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@require_POST
def projeto_delete_ajax(request, pk):
    """Excluir projeto via AJAX"""
    try:
        projeto = get_object_or_404(Projeto, pk=pk)
        projeto.delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# =============================================================================
# 2. COMPRAS
# =============================================================================

# -----------------------------------------------------------------------------
# COTAÇÕES
# -----------------------------------------------------------------------------

@login_required
def cotacao_manager(request):
    """Lista de cotações"""
    queryset = Cotacao.objects.all().order_by('-id')

    search = request.GET.get('search', '')
    if search:
        queryset = queryset.filter(
            Q(numero__icontains=search) | 
            Q(fornecedor__nome__icontains=search)
        )

    status = request.GET.get('status', '')
    if status:
        queryset = queryset.filter(status=status)

    paginator = Paginator(queryset, 25)
    page = request.GET.get('page', 1)
    cotacoes = paginator.get_page(page)

    return render(request, 'compras/cotacao_manager.html', {
        'cotacoes': cotacoes,
        'search': search,
        'status': status
    })


@login_required
def cotacao_form(request, pk=None):
    """Formulário de cotação (ADD/EDIT) - FUNÇÃO ADICIONADA"""
    cotacao = get_object_or_404(Cotacao, pk=pk) if pk else None

    if request.method == 'POST':
        form = CotacaoForm(request.POST, instance=cotacao)
        if form.is_valid():
            cotacao = form.save()
            messages.success(request, f'Cotação {cotacao.numero} salva com sucesso!')
            return redirect('ERP_ServicesBI:cotacao_manager')
    else:
        form = CotacaoForm(instance=cotacao)
        # Se nova cotação, gerar número
        if not cotacao:
            ultima = Cotacao.objects.order_by('-id').first()
            numero = f"COT{((ultima.id + 1) if ultima else 1):06d}"
            form.initial['numero'] = numero

    context = {
        'form': form,
        'cotacao': cotacao,
        'is_edit': bool(pk),
        'fornecedores': Fornecedor.objects.filter(ativo=True),
        'produtos': Produto.objects.filter(ativo=True)
    }
    return render(request, 'compras/cotacao_form.html', context)


@login_required
def cotacao_delete(request, pk):
    """Excluir cotação"""
    cotacao = get_object_or_404(Cotacao, pk=pk)

    if request.method == 'POST':
        cotacao.delete()
        messages.success(request, 'Cotação excluída com sucesso!')
        return redirect('ERP_ServicesBI:cotacao_manager')

    return render(request, 'compras/cotacao_confirm_delete.html', {'cotacao': cotacao})


@login_required
def cotacao_dados_api(request, pk):
    """API para buscar dados da cotação (corrigido - não usa .values())"""
    try:
        cotacao = get_object_or_404(Cotacao, pk=pk)

        itens_list = []
        for item in cotacao.itens.all():
            itens_list.append({
                'id': item.id,
                'produto_id': item.produto_id,
                'produto_codigo': item.produto.codigo if item.produto else '',
                'produto_descricao': item.produto.descricao if item.produto else '',
                'quantidade': float(item.quantidade),
                'preco_unitario': float(item.preco_unitario) if item.preco_unitario else 0,
                'preco_total': float(item.preco_total) if item.preco_total else 0,
                'descricao_display': item.descricao_display,  # @property funciona aqui!
            })

        return JsonResponse({
            'success': True,
            'cotacao': {
                'id': cotacao.id,
                'numero': cotacao.numero,
                'fornecedor_id': cotacao.fornecedor_id,
                'fornecedor_nome': cotacao.fornecedor.nome if cotacao.fornecedor else '',
                'data_emissao': cotacao.data_emissao.isoformat() if cotacao.data_emissao else None,
                'data_validade': cotacao.data_validade.isoformat() if cotacao.data_validade else None,
                'status': cotacao.status,
                'observacoes': cotacao.observacoes or '',
                'valor_total': float(cotacao.valor_total) if cotacao.valor_total else 0,
            },
            'itens': itens_list
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def cotacao_enviar_api(request):
    """API para enviar cotação"""
    try:
        data = json.loads(request.body)
        cotacao_id = data.get('cotacao_id')
        cotacao = get_object_or_404(Cotacao, pk=cotacao_id)

        # Atualizar status
        cotacao.status = 'enviada'
        cotacao.data_envio = timezone.now()
        cotacao.save()

        return JsonResponse({'success': True, 'message': 'Cotação enviada com sucesso!'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def cotacao_item_add_api(request):
    """API para adicionar item à cotação"""
    try:
        data = json.loads(request.body)
        cotacao_id = data.get('cotacao_id')
        cotacao = get_object_or_404(Cotacao, pk=cotacao_id)

        item = ItemCotacao.objects.create(
            cotacao=cotacao,
            produto_id=data.get('produto_id'),
            quantidade=data.get('quantidade'),
            preco_unitario=data.get('preco_unitario', 0),
            observacao=data.get('observacao', '')
        )

        return JsonResponse({
            'success': True, 
            'item_id': item.id,
            'preco_total': float(item.preco_total)
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def cotacao_item_delete_api(request, item_id):
    """API para remover item da cotação"""
    try:
        item = get_object_or_404(ItemCotacao, pk=item_id)
        cotacao_id = item.cotacao_id
        item.delete()

        # Recalcular total
        cotacao = Cotacao.objects.get(pk=cotacao_id)
        cotacao.recalcular_total()

        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# -----------------------------------------------------------------------------
# PEDIDOS DE COMPRA
# -----------------------------------------------------------------------------

@login_required
def pedido_compra_manager(request):
    """Lista de pedidos de compra"""
    queryset = PedidoCompra.objects.all().order_by('-id')

    search = request.GET.get('search', '')
    if search:
        queryset = queryset.filter(
            Q(numero__icontains=search) | 
            Q(fornecedor__nome__icontains=search)
        )

    status = request.GET.get('status', '')
    if status:
        queryset = queryset.filter(status=status)

    paginator = Paginator(queryset, 25)
    page = request.GET.get('page', 1)
    pedidos = paginator.get_page(page)

    return render(request, 'compras/pedido_compra_manager.html', {
        'pedidos': pedidos,
        'search': search,
        'status': status
    })


@login_required
def pedido_compra_form(request, pk=None):
    """Formulário de pedido de compra (ADD/EDIT) - FUNÇÃO ADICIONADA"""
    pedido = get_object_or_404(PedidoCompra, pk=pk) if pk else None

    if request.method == 'POST':
        form = PedidoCompraForm(request.POST, instance=pedido)
        if form.is_valid():
            pedido = form.save()
            messages.success(request, f'Pedido {pedido.numero} salvo com sucesso!')
            return redirect('ERP_ServicesBI:pedido_compra_manager')
    else:
        form = PedidoCompraForm(instance=pedido)
        if not pedido:
            ultimo = PedidoCompra.objects.order_by('-id').first()
            numero = f"PC{((ultimo.id + 1) if ultimo else 1):06d}"
            form.initial['numero'] = numero
            form.initial['data_emissao'] = timezone.now().date()

    context = {
        'form': form,
        'pedido': pedido,
        'is_edit': bool(pk),
        'fornecedores': Fornecedor.objects.filter(ativo=True),
        'produtos': Produto.objects.filter(ativo=True),
        'condicoes': CondicaoPagamento.objects.filter(ativo=True),
        'formas': FormaPagamento.objects.filter(ativo=True)
    }
    return render(request, 'compras/pedido_compra_form.html', context)


@login_required
def pedido_compra_delete(request, pk):
    """Excluir pedido de compra"""
    pedido = get_object_or_404(PedidoCompra, pk=pk)

    if request.method == 'POST':
        pedido.delete()
        messages.success(request, 'Pedido de compra excluído com sucesso!')
        return redirect('ERP_ServicesBI:pedido_compra_manager')

    return render(request, 'compras/pedido_compra_confirm_delete.html', {'pedido': pedido})


@login_required
def pedido_compra_dados_api(request, pk):
    """API para buscar dados do pedido de compra"""
    try:
        pedido = get_object_or_404(PedidoCompra, pk=pk)

        itens_list = []
        for item in pedido.itens.all():
            itens_list.append({
                'id': item.id,
                'produto_id': item.produto_id,
                'produto_codigo': item.produto.codigo if item.produto else '',
                'produto_descricao': item.produto.descricao if item.produto else '',
                'quantidade': float(item.quantidade),
                'quantidade_recebida': float(item.quantidade_recebida) if item.quantidade_recebida else 0,
                'preco_unitario': float(item.preco_unitario),
                'preco_total': float(item.preco_total),
                'data_prevista': item.data_prevista.isoformat() if item.data_prevista else None,
            })

        return JsonResponse({
            'success': True,
            'pedido': {
                'id': pedido.id,
                'numero': pedido.numero,
                'fornecedor_id': pedido.fornecedor_id,
                'fornecedor_nome': pedido.fornecedor.nome if pedido.fornecedor else '',
                'data_emissao': pedido.data_emissao.isoformat() if pedido.data_emissao else None,
                'data_previsao_entrega': pedido.data_previsao_entrega.isoformat() if pedido.data_previsao_entrega else None,
                'status': pedido.status,
                'observacoes': pedido.observacoes or '',
                'valor_total': float(pedido.valor_total) if pedido.valor_total else 0,
                'condicao_pagamento_id': pedido.condicao_pagamento_id,
                'forma_pagamento_id': pedido.forma_pagamento_id,
            },
            'itens': itens_list
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def pedido_salvar_api(request):
    """API para salvar pedido de compra"""
    try:
        data = json.loads(request.body)
        pedido_id = data.get('pedido_id')

        if pedido_id:
            pedido = get_object_or_404(PedidoCompra, pk=pedido_id)
        else:
            pedido = PedidoCompra()

        pedido.fornecedor_id = data.get('fornecedor_id')
        pedido.data_emissao = data.get('data_emissao')
        pedido.data_previsao_entrega = data.get('data_previsao_entrega')
        pedido.condicao_pagamento_id = data.get('condicao_pagamento_id')
        pedido.forma_pagamento_id = data.get('forma_pagamento_id')
        pedido.observacoes = data.get('observacoes', '')
        pedido.save()

        # Processar itens
        itens_data = data.get('itens', [])
        for item_data in itens_data:
            if item_data.get('id'):
                item = get_object_or_404(ItemPedidoCompra, pk=item_data['id'])
            else:
                item = ItemPedidoCompra(pedido=pedido)

            item.produto_id = item_data.get('produto_id')
            item.quantidade = item_data.get('quantidade')
            item.preco_unitario = item_data.get('preco_unitario')
            item.data_prevista = item_data.get('data_prevista')
            item.save()

        pedido.recalcular_total()

        return JsonResponse({'success': True, 'pedido_id': pedido.id})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def pedido_item_add_api(request):
    """API para adicionar item ao pedido"""
    try:
        data = json.loads(request.body)
        pedido_id = data.get('pedido_id')
        pedido = get_object_or_404(PedidoCompra, pk=pedido_id)

        item = ItemPedidoCompra.objects.create(
            pedido=pedido,
            produto_id=data.get('produto_id'),
            quantidade=data.get('quantidade'),
            preco_unitario=data.get('preco_unitario', 0),
            data_prevista=data.get('data_prevista'),
            observacao=data.get('observacao', '')
        )

        pedido.recalcular_total()

        return JsonResponse({
            'success': True, 
            'item_id': item.id,
            'preco_total': float(item.preco_total)
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def pedido_item_delete_api(request, item_id):
    """API para remover item do pedido"""
    try:
        item = get_object_or_404(ItemPedidoCompra, pk=item_id)
        pedido = item.pedido
        item.delete()
        pedido.recalcular_total()

        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def api_enviar_aprovacao(request):
    """API para enviar pedido para aprovação - FUNÇÃO ADICIONADA"""
    try:
        data = json.loads(request.body)
        pedido_id = data.get('pedido_id')
        pedido = get_object_or_404(PedidoCompra, pk=pedido_id)

        # Verificar se já está em aprovação
        if pedido.status == 'em_aprovacao':
            return JsonResponse({
                'success': False, 
                'error': 'Pedido já está em processo de aprovação'
            })

        # Criar registro de aprovação
        PedidoAprovacao.objects.create(
            pedido=pedido,
            nivel=1,
            status='pendente',
            solicitante=request.user,
            data_solicitacao=timezone.now()
        )

        # Atualizar status do pedido
        pedido.status = 'em_aprovacao'
        pedido.save()

        return JsonResponse({
            'success': True, 
            'message': 'Pedido enviado para aprovação com sucesso!'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# -----------------------------------------------------------------------------
# NOTAS FISCAIS DE ENTRADA
# -----------------------------------------------------------------------------

@login_required
def nota_fiscal_entrada_manager(request):
    """Lista de notas fiscais de entrada"""
    queryset = NotaFiscalEntrada.objects.all().order_by('-id')

    search = request.GET.get('search', '')
    if search:
        queryset = queryset.filter(
            Q(numero__icontains=search) | 
            Q(fornecedor__nome__icontains=search) |
            Q(chave_acesso__icontains=search)
        )

    status = request.GET.get('status', '')
    if status:
        queryset = queryset.filter(status=status)

    paginator = Paginator(queryset, 25)
    page = request.GET.get('page', 1)
    notas = paginator.get_page(page)

    return render(request, 'compras/nota_fiscal_entrada_manager.html', {
        'notas': notas,
        'search': search,
        'status': status
    })


@login_required
def nota_fiscal_entrada_form(request, pk=None):
    """Formulário de nota fiscal de entrada (ADD/EDIT) - FUNÇÃO ADICIONADA"""
    nota = get_object_or_404(NotaFiscalEntrada, pk=pk) if pk else None

    if request.method == 'POST':
        form = NotaFiscalEntradaForm(request.POST, instance=nota)
        if form.is_valid():
            nota = form.save()
            messages.success(request, f'Nota Fiscal {nota.numero} salva com sucesso!')
            return redirect('ERP_ServicesBI:nota_fiscal_entrada_manager')
    else:
        form = NotaFiscalEntradaForm(instance=nota)
        if not nota:
            form.initial['data_recebimento'] = timezone.now().date()  # CORREÇÃO: data_recebimento

    context = {
        'form': form,
        'nota': nota,
        'is_edit': bool(pk),
        'fornecedores': Fornecedor.objects.filter(ativo=True),
        'produtos': Produto.objects.filter(ativo=True),
        'depositos': Deposito.objects.filter(ativo=True)
    }
    return render(request, 'compras/nota_fiscal_entrada_form.html', context)


@login_required
def nota_fiscal_entrada_delete(request, pk):
    """Excluir nota fiscal de entrada"""
    nota = get_object_or_404(NotaFiscalEntrada, pk=pk)

    if request.method == 'POST':
        nota.delete()
        messages.success(request, 'Nota Fiscal excluída com sucesso!')
        return redirect('ERP_ServicesBI:nota_fiscal_entrada_manager')

    return render(request, 'compras/nota_fiscal_entrada_confirm_delete.html', {'nota': nota})


@login_required
def entrada_nfe(request):
    """Tela de entrada de NFe"""
    return render(request, 'compras/entrada_nfe.html', {
        'fornecedores': Fornecedor.objects.filter(ativo=True),
        'produtos': Produto.objects.filter(ativo=True),
        'depositos': Deposito.objects.filter(ativo=True)
    })


@login_required
@require_POST
def nota_fiscal_salvar_api(request):
    """API para salvar nota fiscal de entrada"""
    try:
        data = json.loads(request.body)
        nota_id = data.get('nota_id')

        if nota_id:
            nota = get_object_or_404(NotaFiscalEntrada, pk=nota_id)
        else:
            nota = NotaFiscalEntrada()

        nota.fornecedor_id = data.get('fornecedor_id')
        nota.numero = data.get('numero')
        nota.serie = data.get('serie', '1')
        nota.chave_acesso = data.get('chave_acesso', '')
        nota.data_emissao = data.get('data_emissao')
        nota.data_recebimento = data.get('data_recebimento')  # CORREÇÃO: data_recebimento
        nota.deposito_id = data.get('deposito_id')
        nota.observacoes = data.get('observacoes', '')
        nota.save()

        # Processar itens
        itens_data = data.get('itens', [])
        for item_data in itens_data:
            if item_data.get('id'):
                item = get_object_or_404(ItemNotaFiscalEntrada, pk=item_data['id'])
            else:
                item = ItemNotaFiscalEntrada(nota_fiscal=nota)

            item.produto_id = item_data.get('produto_id')
            item.quantidade = item_data.get('quantidade')
            item.preco_unitario = item_data.get('preco_unitario')
            item.cfop = item_data.get('cfop', '')
            item.ncm = item_data.get('ncm', '')
            item.save()

        nota.recalcular_total()

        return JsonResponse({'success': True, 'nota_id': nota.id})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def nota_fiscal_entrada_dados_api(request, pk):
    """API para buscar dados da nota fiscal de entrada"""
    try:
        nota = get_object_or_404(NotaFiscalEntrada, pk=pk)

        itens_list = []
        for item in nota.itens.all():
            itens_list.append({
                'id': item.id,
                'produto_id': item.produto_id,
                'produto_codigo': item.produto.codigo if item.produto else '',
                'produto_descricao': item.produto.descricao if item.produto else '',
                'quantidade': float(item.quantidade),
                'preco_unitario': float(item.preco_unitario),
                'preco_total': float(item.preco_total),
                'cfop': item.cfop or '',
                'ncm': item.ncm or '',
            })

        return JsonResponse({
            'success': True,
            'nota': {
                'id': nota.id,
                'numero': nota.numero,
                'serie': nota.serie,
                'chave_acesso': nota.chave_acesso or '',
                'fornecedor_id': nota.fornecedor_id,
                'fornecedor_nome': nota.fornecedor.nome if nota.fornecedor else '',
                'data_emissao': nota.data_emissao.isoformat() if nota.data_emissao else None,
                'data_recebimento': nota.data_recebimento.isoformat() if nota.data_recebimento else None,  # CORREÇÃO
                'deposito_id': nota.deposito_id,
                'deposito_nome': nota.deposito.nome if nota.deposito else '',
                'status': nota.status,
                'observacoes': nota.observacoes or '',
                'valor_total': float(nota.valor_total) if nota.valor_total else 0,
                'valor_icms': float(nota.valor_icms) if nota.valor_icms else 0,
                'valor_ipi': float(nota.valor_ipi) if nota.valor_ipi else 0,
            },
            'itens': itens_list
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def nota_fiscal_entrada_item_add_api(request):
    """API para adicionar item à NF entrada"""
    try:
        data = json.loads(request.body)
        nota_id = data.get('nota_id')
        nota = get_object_or_404(NotaFiscalEntrada, pk=nota_id)

        item = ItemNotaFiscalEntrada.objects.create(
            nota_fiscal=nota,
            produto_id=data.get('produto_id'),
            quantidade=data.get('quantidade'),
            preco_unitario=data.get('preco_unitario', 0),
            cfop=data.get('cfop', ''),
            ncm=data.get('ncm', '')
        )

        nota.recalcular_total()

        return JsonResponse({
            'success': True, 
            'item_id': item.id,
            'preco_total': float(item.preco_total)
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def nota_fiscal_entrada_item_delete_api(request, item_id):
    """API para remover item da NF entrada"""
    try:
        item = get_object_or_404(ItemNotaFiscalEntrada, pk=item_id)
        nota = item.nota_fiscal
        item.delete()
        nota.recalcular_total()

        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def nota_fiscal_entrada_confirmar_recebimento(request, pk):
    """Confirmar recebimento da NF e gerar movimentação de estoque"""
    try:
        nota = get_object_or_404(NotaFiscalEntrada, pk=pk)

        with transaction.atomic():
            # Criar movimentação de estoque de entrada
            movimentacao = MovimentacaoEstoque.objects.create(
                tipo='entrada',
                data=timezone.now(),
                deposito=nota.deposito,
                nota_fiscal_entrada=nota,
                observacao=f'Entrada via NF {nota.numero} - {nota.fornecedor.nome if nota.fornecedor else ""}'
            )

            # Criar itens da movimentação
            for item_nota in nota.itens.all():
                ItemMovimentacaoEstoque.objects.create(
                    movimentacao=movimentacao,
                    produto=item_nota.produto,
                    quantidade=item_nota.quantidade,
                    preco_unitario=item_nota.preco_unitario
                )

                # Atualizar saldo do produto no depósito
                posicao, created = PosicaoEstoque.objects.get_or_create(
                    produto=item_nota.produto,
                    deposito=nota.deposito,
                    defaults={'quantidade': 0}
                )
                posicao.quantidade += item_nota.quantidade
                posicao.save()

            # Atualizar status da nota
            nota.status = 'recebida'
            nota.save()

            # Atualizar pedido relacionado se houver
            if nota.pedido_compra:
                nota.pedido_compra.atualizar_status_recebimento()

        return JsonResponse({
            'success': True, 
            'message': 'Recebimento confirmado e estoque atualizado!'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# -----------------------------------------------------------------------------
# RELATÓRIO DE COMPRAS
# -----------------------------------------------------------------------------

@login_required
def relatorio_compras(request):
    """Relatório de compras"""
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    fornecedor_id = request.GET.get('fornecedor_id')

    queryset = NotaFiscalEntrada.objects.filter(status='recebida')

    if data_inicio:
        queryset = queryset.filter(data_recebimento__gte=data_inicio)  # CORREÇÃO
    if data_fim:
        queryset = queryset.filter(data_recebimento__lte=data_fim)  # CORREÇÃO
    if fornecedor_id:
        queryset = queryset.filter(fornecedor_id=fornecedor_id)

    total_compras = queryset.aggregate(total=Sum('valor_total'))['total'] or 0

    # Agrupar por fornecedor
    por_fornecedor = queryset.values('fornecedor__nome').annotate(
        total=Sum('valor_total'),
        quantidade=Count('id')
    ).order_by('-total')

    # Agrupar por mês
    por_mes = queryset.extra(
        select={'mes': "TO_CHAR(data_recebimento, 'YYYY-MM')"}  # CORREÇÃO
    ).values('mes').annotate(
        total=Sum('valor_total'),
        quantidade=Count('id')
    ).order_by('mes')

    return render(request, 'compras/relatorio_compras.html', {
        'notas': queryset,
        'total_compras': total_compras,
        'por_fornecedor': por_fornecedor,
        'por_mes': por_mes,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'fornecedor_id': fornecedor_id,
        'fornecedores': Fornecedor.objects.filter(ativo=True)
    })


# =============================================================================
# 3. VENDAS
# =============================================================================

# -----------------------------------------------------------------------------
# ORÇAMENTOS
# -----------------------------------------------------------------------------

@login_required
def orcamento_manager(request):
    """Lista de orçamentos"""
    queryset = Orcamento.objects.all().order_by('-id')

    search = request.GET.get('search', '')
    if search:
        queryset = queryset.filter(
            Q(numero__icontains=search) | 
            Q(cliente__nome__icontains=search)
        )

    status = request.GET.get('status', '')
    if status:
        queryset = queryset.filter(status=status)

    paginator = Paginator(queryset, 25)
    page = request.GET.get('page', 1)
    orcamentos = paginator.get_page(page)

    return render(request, 'vendas/orcamento_manager.html', {
        'orcamentos': orcamentos,
        'search': search,
        'status': status
    })


@login_required
def orcamento_form(request, pk=None):
    """Formulário de orçamento (add/edit)"""
    orcamento = get_object_or_404(Orcamento, pk=pk) if pk else None

    if request.method == 'POST':
        form = OrcamentoForm(request.POST, instance=orcamento)
        if form.is_valid():
            orcamento = form.save()
            messages.success(request, f'Orçamento {orcamento.numero} salvo com sucesso!')
            return redirect('ERP_ServicesBI:orcamento_manager')
    else:
        form = OrcamentoForm(instance=orcamento)
        if not orcamento:
            ultimo = Orcamento.objects.order_by('-id').first()
            numero = f"ORC{((ultimo.id + 1) if ultimo else 1):06d}"
            form.initial['numero'] = numero
            form.initial['data_emissao'] = timezone.now().date()

    context = {
        'form': form,
        'orcamento': orcamento,
        'is_edit': bool(pk),
        'clientes': Cliente.objects.filter(ativo=True),
        'vendedores': Vendedor.objects.filter(ativo=True),
        'produtos': Produto.objects.filter(ativo=True),
        'condicoes': CondicaoPagamento.objects.filter(ativo=True),
        'formas': FormaPagamento.objects.filter(ativo=True)
    }
    return render(request, 'vendas/orcamento_form.html', context)


@login_required
def orcamento_delete(request, pk):
    """Excluir orçamento"""
    orcamento = get_object_or_404(Orcamento, pk=pk)

    if request.method == 'POST':
        orcamento.delete()
        messages.success(request, 'Orçamento excluído com sucesso!')
        return redirect('ERP_ServicesBI:orcamento_manager')

    return render(request, 'vendas/orcamento_confirm_delete.html', {'orcamento': orcamento})


@login_required
def orcamento_dados_api(request, pk):
    """API para buscar dados do orçamento"""
    try:
        orcamento = get_object_or_404(Orcamento, pk=pk)

        itens_list = []
        for item in orcamento.itens.all():
            itens_list.append({
                'id': item.id,
                'produto_id': item.produto_id,
                'produto_codigo': item.produto.codigo if item.produto else '',
                'produto_descricao': item.produto.descricao if item.produto else '',
                'quantidade': float(item.quantidade),
                'preco_unitario': float(item.preco_unitario),
                'desconto_percentual': float(item.desconto_percentual) if item.desconto_percentual else 0,
                'preco_total': float(item.preco_total),
            })

        return JsonResponse({
            'success': True,
            'orcamento': {
                'id': orcamento.id,
                'numero': orcamento.numero,
                'cliente_id': orcamento.cliente_id,
                'cliente_nome': orcamento.cliente.nome if orcamento.cliente else '',
                'vendedor_id': orcamento.vendedor_id,
                'data_emissao': orcamento.data_emissao.isoformat() if orcamento.data_emissao else None,
                'data_validade': orcamento.data_validade.isoformat() if orcamento.data_validade else None,
                'status': orcamento.status,
                'observacoes': orcamento.observacoes or '',
                'valor_total': float(orcamento.valor_total) if orcamento.valor_total else 0,
                'valor_desconto': float(orcamento.valor_desconto) if orcamento.valor_desconto else 0,
                'condicao_pagamento_id': orcamento.condicao_pagamento_id,
                'forma_pagamento_id': orcamento.forma_pagamento_id,
            },
            'itens': itens_list
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def orcamento_salvar_api(request):
    """API para salvar orçamento"""
    try:
        data = json.loads(request.body)
        orcamento_id = data.get('orcamento_id')

        if orcamento_id:
            orcamento = get_object_or_404(Orcamento, pk=orcamento_id)
        else:
            orcamento = Orcamento()

        orcamento.cliente_id = data.get('cliente_id')
        orcamento.vendedor_id = data.get('vendedor_id')
        orcamento.data_emissao = data.get('data_emissao')
        orcamento.data_validade = data.get('data_validade')
        orcamento.condicao_pagamento_id = data.get('condicao_pagamento_id')
        orcamento.forma_pagamento_id = data.get('forma_pagamento_id')
        orcamento.observacoes = data.get('observacoes', '')
        orcamento.save()

        # Processar itens
        itens_data = data.get('itens', [])
        for item_data in itens_data:
            if item_data.get('id'):
                item = get_object_or_404(ItemOrcamento, pk=item_data['id'])
            else:
                item = ItemOrcamento(orcamento=orcamento)

            item.produto_id = item_data.get('produto_id')
            item.quantidade = item_data.get('quantidade')
            item.preco_unitario = item_data.get('preco_unitario')
            item.desconto_percentual = item_data.get('desconto_percentual', 0)
            item.save()

        orcamento.recalcular_total()

        return JsonResponse({'success': True, 'orcamento_id': orcamento.id})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def orcamento_item_add_api(request):
    """API para adicionar item ao orçamento"""
    try:
        data = json.loads(request.body)
        orcamento_id = data.get('orcamento_id')
        orcamento = get_object_or_404(Orcamento, pk=orcamento_id)

        item = ItemOrcamento.objects.create(
            orcamento=orcamento,
            produto_id=data.get('produto_id'),
            quantidade=data.get('quantidade'),
            preco_unitario=data.get('preco_unitario', 0),
            desconto_percentual=data.get('desconto_percentual', 0)
        )

        orcamento.recalcular_total()

        return JsonResponse({
            'success': True, 
            'item_id': item.id,
            'preco_total': float(item.preco_total)
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def orcamento_item_delete_api(request, item_id):
    """API para remover item do orçamento"""
    try:
        item = get_object_or_404(ItemOrcamento, pk=item_id)
        orcamento = item.orcamento
        item.delete()
        orcamento.recalcular_total()

        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def orcamento_gerar_pedido(request, pk):
    """Gerar pedido de venda a partir do orçamento"""
    try:
        orcamento = get_object_or_404(Orcamento, pk=pk)

        with transaction.atomic():
            # Criar pedido
            ultimo = PedidoVenda.objects.order_by('-id').first()
            numero = f"PV{((ultimo.id + 1) if ultimo else 1):06d}"

            pedido = PedidoVenda.objects.create(
                numero=numero,
                cliente=orcamento.cliente,
                vendedor=orcamento.vendedor,
                orcamento=orcamento,
                data_emissao=timezone.now().date(),
                condicao_pagamento=orcamento.condicao_pagamento,
                forma_pagamento=orcamento.forma_pagamento,
                observacoes=f'Gerado do orçamento {orcamento.numero}',
                valor_total=orcamento.valor_total,
                status='aberto'
            )

            # Copiar itens
            for item_orc in orcamento.itens.all():
                ItemPedidoVenda.objects.create(
                    pedido=pedido,
                    produto=item_orc.produto,
                    quantidade=item_orc.quantidade,
                    preco_unitario=item_orc.preco_unitario,
                    desconto_percentual=item_orc.desconto_percentual,
                    preco_total=item_orc.preco_total
                )

            # Atualizar status do orçamento
            orcamento.status = 'aprovado'
            orcamento.save()

        return JsonResponse({
            'success': True, 
            'pedido_id': pedido.id,
            'pedido_numero': pedido.numero,
            'message': f'Pedido {pedido.numero} gerado com sucesso!'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# -----------------------------------------------------------------------------
# ORÇAMENTOS DE PROJETO
# -----------------------------------------------------------------------------

@login_required
def orcamento_projeto_manager(request):
    """Lista de orçamentos de projeto"""
    queryset = OrcamentoProjeto.objects.all().order_by('-id')

    search = request.GET.get('search', '')
    if search:
        queryset = queryset.filter(
            Q(numero__icontains=search) | 
            Q(projeto__nome__icontains=search)
        )

    status = request.GET.get('status', '')
    if status:
        queryset = queryset.filter(status=status)

    paginator = Paginator(queryset, 25)
    page = request.GET.get('page', 1)
    orcamentos = paginator.get_page(page)

    return render(request, 'vendas/orcamento_projeto_manager.html', {
        'orcamentos': orcamentos,
        'search': search,
        'status': status
    })


@login_required
def orcamento_projeto_form(request, pk=None):
    """Formulário de orçamento de projeto"""
    orcamento = get_object_or_404(OrcamentoProjeto, pk=pk) if pk else None

    if request.method == 'POST':
        form = OrcamentoProjetoForm(request.POST, instance=orcamento)
        if form.is_valid():
            orcamento = form.save()
            messages.success(request, f'Orçamento de projeto {orcamento.numero} salvo com sucesso!')
            return redirect('ERP_ServicesBI:orcamento_projeto_manager')
    else:
        form = OrcamentoProjetoForm(instance=orcamento)
        if not orcamento:
            ultimo = OrcamentoProjeto.objects.order_by('-id').first()
            numero = f"ORCP{((ultimo.id + 1) if ultimo else 1):06d}"
            form.initial['numero'] = numero

    context = {
        'form': form,
        'orcamento': orcamento,
        'is_edit': bool(pk),
        'projetos': Projeto.objects.filter(status='ativo'),
        'clientes': Cliente.objects.filter(ativo=True)
    }
    return render(request, 'vendas/orcamento_projeto_form.html', context)


@login_required
def orcamento_projeto_delete(request, pk):
    """Excluir orçamento de projeto"""
    orcamento = get_object_or_404(OrcamentoProjeto, pk=pk)

    if request.method == 'POST':
        orcamento.delete()
        messages.success(request, 'Orçamento de projeto excluído com sucesso!')
        return redirect('ERP_ServicesBI:orcamento_projeto_manager')

    return render(request, 'vendas/orcamento_projeto_confirm_delete.html', {'orcamento': orcamento})


# -----------------------------------------------------------------------------
# PEDIDOS DE VENDA
# -----------------------------------------------------------------------------

@login_required
def pedido_venda_manager(request):
    """Lista de pedidos de venda"""
    queryset = PedidoVenda.objects.all().order_by('-id')

    search = request.GET.get('search', '')
    if search:
        queryset = queryset.filter(
            Q(numero__icontains=search) | 
            Q(cliente__nome__icontains=search)
        )

    status = request.GET.get('status', '')
    if status:
        queryset = queryset.filter(status=status)

    paginator = Paginator(queryset, 25)
    page = request.GET.get('page', 1)
    pedidos = paginator.get_page(page)

    return render(request, 'vendas/pedido_venda_manager.html', {
        'pedidos': pedidos,
        'search': search,
        'status': status
    })


@login_required
def pedido_venda_form(request, pk=None):
    """Formulário de pedido de venda"""
    pedido = get_object_or_404(PedidoVenda, pk=pk) if pk else None

    if request.method == 'POST':
        form = PedidoVendaForm(request.POST, instance=pedido)
        if form.is_valid():
            pedido = form.save()
            messages.success(request, f'Pedido {pedido.numero} salvo com sucesso!')
            return redirect('ERP_ServicesBI:pedido_venda_manager')
    else:
        form = PedidoVendaForm(instance=pedido)
        if not pedido:
            ultimo = PedidoVenda.objects.order_by('-id').first()
            numero = f"PV{((ultimo.id + 1) if ultimo else 1):06d}"
            form.initial['numero'] = numero
            form.initial['data_emissao'] = timezone.now().date()

    context = {
        'form': form,
        'pedido': pedido,
        'is_edit': bool(pk),
        'clientes': Cliente.objects.filter(ativo=True),
        'vendedores': Vendedor.objects.filter(ativo=True),
        'produtos': Produto.objects.filter(ativo=True),
        'condicoes': CondicaoPagamento.objects.filter(ativo=True),
        'formas': FormaPagamento.objects.filter(ativo=True)
    }
    return render(request, 'vendas/pedido_venda_form.html', context)


@login_required
def pedido_venda_delete(request, pk):
    """Excluir pedido de venda"""
    pedido = get_object_or_404(PedidoVenda, pk=pk)

    if request.method == 'POST':
        pedido.delete()
        messages.success(request, 'Pedido de venda excluído com sucesso!')
        return redirect('ERP_ServicesBI:pedido_venda_manager')

    return render(request, 'vendas/pedido_venda_confirm_delete.html', {'pedido': pedido})


@login_required
def pedido_venda_dados_api(request, pk):
    """API para buscar dados do pedido de venda"""
    try:
        pedido = get_object_or_404(PedidoVenda, pk=pk)

        itens_list = []
        for item in pedido.itens.all():
            itens_list.append({
                'id': item.id,
                'produto_id': item.produto_id,
                'produto_codigo': item.produto.codigo if item.produto else '',
                'produto_descricao': item.produto.descricao if item.produto else '',
                'quantidade': float(item.quantidade),
                'quantidade_atendida': float(item.quantidade_atendida) if item.quantidade_atendida else 0,
                'preco_unitario': float(item.preco_unitario),
                'desconto_percentual': float(item.desconto_percentual) if item.desconto_percentual else 0,
                'preco_total': float(item.preco_total),
            })

        return JsonResponse({
            'success': True,
            'pedido': {
                'id': pedido.id,
                'numero': pedido.numero,
                'cliente_id': pedido.cliente_id,
                'cliente_nome': pedido.cliente.nome if pedido.cliente else '',
                'vendedor_id': pedido.vendedor_id,
                'data_emissao': pedido.data_emissao.isoformat() if pedido.data_emissao else None,
                'status': pedido.status,
                'observacoes': pedido.observacoes or '',
                'valor_total': float(pedido.valor_total) if pedido.valor_total else 0,
                'condicao_pagamento_id': pedido.condicao_pagamento_id,
                'forma_pagamento_id': pedido.forma_pagamento_id,
            },
            'itens': itens_list
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def pedido_venda_salvar_api(request):
    """API para salvar pedido de venda"""
    try:
        data = json.loads(request.body)
        pedido_id = data.get('pedido_id')

        if pedido_id:
            pedido = get_object_or_404(PedidoVenda, pk=pedido_id)
        else:
            pedido = PedidoVenda()

        pedido.cliente_id = data.get('cliente_id')
        pedido.vendedor_id = data.get('vendedor_id')
        pedido.data_emissao = data.get('data_emissao')
        pedido.condicao_pagamento_id = data.get('condicao_pagamento_id')
        pedido.forma_pagamento_id = data.get('forma_pagamento_id')
        pedido.observacoes = data.get('observacoes', '')
        pedido.save()

        # Processar itens
        itens_data = data.get('itens', [])
        for item_data in itens_data:
            if item_data.get('id'):
                item = get_object_or_404(ItemPedidoVenda, pk=item_data['id'])
            else:
                item = ItemPedidoVenda(pedido=pedido)

            item.produto_id = item_data.get('produto_id')
            item.quantidade = item_data.get('quantidade')
            item.preco_unitario = item_data.get('preco_unitario')
            item.desconto_percentual = item_data.get('desconto_percentual', 0)
            item.save()

        pedido.recalcular_total()

        return JsonResponse({'success': True, 'pedido_id': pedido.id})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def pedido_venda_item_add_api(request):
    """API para adicionar item ao pedido de venda"""
    try:
        data = json.loads(request.body)
        pedido_id = data.get('pedido_id')
        pedido = get_object_or_404(PedidoVenda, pk=pedido_id)

        item = ItemPedidoVenda.objects.create(
            pedido=pedido,
            produto_id=data.get('produto_id'),
            quantidade=data.get('quantidade'),
            preco_unitario=data.get('preco_unitario', 0),
            desconto_percentual=data.get('desconto_percentual', 0)
        )

        pedido.recalcular_total()

        return JsonResponse({
            'success': True, 
            'item_id': item.id,
            'preco_total': float(item.preco_total)
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def pedido_venda_item_delete_api(request, item_id):
    """API para remover item do pedido de venda"""
    try:
        item = get_object_or_404(ItemPedidoVenda, pk=item_id)
        pedido = item.pedido
        item.delete()
        pedido.recalcular_total()

        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def pedido_venda_gerar_nfe(request, pk):
    """Gerar nota fiscal de saída a partir do pedido"""
    try:
        pedido = get_object_or_404(PedidoVenda, pk=pk)

        with transaction.atomic():
            # Criar NF
            ultima = NotaFiscalSaida.objects.order_by('-id').first()
            numero = f"NFS{((ultima.id + 1) if ultima else 1):06d}"

            nota = NotaFiscalSaida.objects.create(
                numero=numero,
                serie='1',
                cliente=pedido.cliente,
                pedido=pedido,
                data_emissao=timezone.now().date(),
                condicao_pagamento=pedido.condicao_pagamento,
                forma_pagamento=pedido.forma_pagamento,
                observacoes=f'Gerado do pedido {pedido.numero}',
                valor_total=0,
                status='emitida'
            )

            # Copiar itens
            total = 0
            for item_ped in pedido.itens.all():
                # Verificar saldo
                saldo = PosicaoEstoque.objects.filter(
                    produto=item_ped.produto
                ).aggregate(s=Sum('quantidade'))['s'] or 0

                if saldo < item_ped.quantidade:
                    raise Exception(f'Saldo insuficiente para o produto {item_ped.produto.codigo}')

                item_nf = ItemNotaFiscalSaida.objects.create(
                    nota_fiscal=nota,
                    produto=item_ped.produto,
                    quantidade=item_ped.quantidade,
                    preco_unitario=item_ped.preco_unitario,
                    desconto_percentual=item_ped.desconto_percentual,
                    preco_total=item_ped.preco_total
                )
                total += item_nf.preco_total

                # Atualizar quantidade atendida no pedido
                item_ped.quantidade_atendida = item_ped.quantidade
                item_ped.save()

            nota.valor_total = total
            nota.save()

            # Atualizar status do pedido
            pedido.status = 'faturado'
            pedido.save()

        return JsonResponse({
            'success': True, 
            'nota_id': nota.id,
            'nota_numero': nota.numero,
            'message': f'Nota Fiscal {nota.numero} gerada com sucesso!'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# -----------------------------------------------------------------------------
# NOTAS FISCAIS DE SAÍDA
# -----------------------------------------------------------------------------

@login_required
def nota_fiscal_saida_manager(request):
    """Lista de notas fiscais de saída"""
    queryset = NotaFiscalSaida.objects.all().order_by('-id')

    search = request.GET.get('search', '')
    if search:
        queryset = queryset.filter(
            Q(numero__icontains=search) | 
            Q(cliente__nome__icontains=search)
        )

    status = request.GET.get('status', '')
    if status:
        queryset = queryset.filter(status=status)

    paginator = Paginator(queryset, 25)
    page = request.GET.get('page', 1)
    notas = paginator.get_page(page)

    return render(request, 'vendas/nota_fiscal_saida_manager.html', {
        'notas': notas,
        'search': search,
        'status': status
    })


@login_required
def nota_fiscal_saida_form(request, pk=None):
    """Formulário de nota fiscal de saída"""
    nota = get_object_or_404(NotaFiscalSaida, pk=pk) if pk else None

    if request.method == 'POST':
        form = NotaFiscalSaidaForm(request.POST, instance=nota)
        if form.is_valid():
            nota = form.save()
            messages.success(request, f'Nota Fiscal {nota.numero} salva com sucesso!')
            return redirect('ERP_ServicesBI:nota_fiscal_saida_manager')
    else:
        form = NotaFiscalSaidaForm(instance=nota)
        if not nota:
            form.initial['data_emissao'] = timezone.now().date()

    context = {
        'form': form,
        'nota': nota,
        'is_edit': bool(pk),
        'clientes': Cliente.objects.filter(ativo=True),
        'produtos': Produto.objects.filter(ativo=True)
    }
    return render(request, 'vendas/nota_fiscal_saida_form.html', context)


@login_required
def nota_fiscal_saida_delete(request, pk):
    """Excluir nota fiscal de saída"""
    nota = get_object_or_404(NotaFiscalSaida, pk=pk)

    if request.method == 'POST':
        nota.delete()
        messages.success(request, 'Nota Fiscal excluída com sucesso!')
        return redirect('ERP_ServicesBI:nota_fiscal_saida_manager')

    return render(request, 'vendas/nota_fiscal_saida_confirm_delete.html', {'nota': nota})


@login_required
def nota_fiscal_saida_dados_api(request, pk):
    """API para buscar dados da nota fiscal de saída"""
    try:
        nota = get_object_or_404(NotaFiscalSaida, pk=pk)

        itens_list = []
        for item in nota.itens.all():
            itens_list.append({
                'id': item.id,
                'produto_id': item.produto_id,
                'produto_codigo': item.produto.codigo if item.produto else '',
                'produto_descricao': item.produto.descricao if item.produto else '',
                'quantidade': float(item.quantidade),
                'preco_unitario': float(item.preco_unitario),
                'desconto_percentual': float(item.desconto_percentual) if item.desconto_percentual else 0,
                'preco_total': float(item.preco_total),
                'cfop': item.cfop or '',
                'ncm': item.ncm or '',
            })

        return JsonResponse({
            'success': True,
            'nota': {
                'id': nota.id,
                'numero': nota.numero,
                'serie': nota.serie,
                'cliente_id': nota.cliente_id,
                'cliente_nome': nota.cliente.nome if nota.cliente else '',
                'data_emissao': nota.data_emissao.isoformat() if nota.data_emissao else None,
                'status': nota.status,
                'observacoes': nota.observacoes or '',
                'valor_total': float(nota.valor_total) if nota.valor_total else 0,
                'valor_icms': float(nota.valor_icms) if nota.valor_icms else 0,
                'valor_ipi': float(nota.valor_ipi) if nota.valor_ipi else 0,
            },
            'itens': itens_list
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def nota_fiscal_saida_salvar_api(request):
    """API para salvar nota fiscal de saída"""
    try:
        data = json.loads(request.body)
        nota_id = data.get('nota_id')

        if nota_id:
            nota = get_object_or_404(NotaFiscalSaida, pk=nota_id)
        else:
            nota = NotaFiscalSaida()

        nota.cliente_id = data.get('cliente_id')
        nota.numero = data.get('numero')
        nota.serie = data.get('serie', '1')
        nota.data_emissao = data.get('data_emissao')
        nota.observacoes = data.get('observacoes', '')
        nota.save()

        # Processar itens
        itens_data = data.get('itens', [])
        for item_data in itens_data:
            if item_data.get('id'):
                item = get_object_or_404(ItemNotaFiscalSaida, pk=item_data['id'])
            else:
                item = ItemNotaFiscalSaida(nota_fiscal=nota)

            item.produto_id = item_data.get('produto_id')
            item.quantidade = item_data.get('quantidade')
            item.preco_unitario = item_data.get('preco_unitario')
            item.desconto_percentual = item_data.get('desconto_percentual', 0)
            item.cfop = item_data.get('cfop', '')
            item.ncm = item_data.get('ncm', '')
            item.save()

        nota.recalcular_total()

        return JsonResponse({'success': True, 'nota_id': nota.id})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def nota_fiscal_saida_item_add_api(request):
    """API para adicionar item à NF saída"""
    try:
        data = json.loads(request.body)
        nota_id = data.get('nota_id')
        nota = get_object_or_404(NotaFiscalSaida, pk=nota_id)

        item = ItemNotaFiscalSaida.objects.create(
            nota_fiscal=nota,
            produto_id=data.get('produto_id'),
            quantidade=data.get('quantidade'),
            preco_unitario=data.get('preco_unitario', 0),
            desconto_percentual=data.get('desconto_percentual', 0),
            cfop=data.get('cfop', ''),
            ncm=data.get('ncm', '')
        )

        nota.recalcular_total()

        return JsonResponse({
            'success': True, 
            'item_id': item.id,
            'preco_total': float(item.preco_total)
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def nota_fiscal_saida_item_delete_api(request, item_id):
    """API para remover item da NF saída"""
    try:
        item = get_object_or_404(ItemNotaFiscalSaida, pk=item_id)
        nota = item.nota_fiscal
        item.delete()
        nota.recalcular_total()

        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def nota_fiscal_saida_confirmar_entrega(request, pk):
    """Confirmar entrega da NF saída e gerar movimentação de estoque"""
    try:
        nota = get_object_or_404(NotaFiscalSaida, pk=pk)

        with transaction.atomic():
            # Criar movimentação de estoque de saída
            deposito = Deposito.objects.filter(ativo=True).first()
            if not deposito:
                raise Exception('Nenhum depósito ativo encontrado')

            movimentacao = MovimentacaoEstoque.objects.create(
                tipo='saida',
                data=timezone.now(),
                deposito=deposito,
                nota_fiscal_saida=nota,
                observacao=f'Saída via NF {nota.numero} - {nota.cliente.nome if nota.cliente else ""}'
            )

            # Criar itens da movimentação e verificar saldo
            for item_nota in nota.itens.all():
                # Verificar saldo
                saldo = PosicaoEstoque.objects.filter(
                    produto=item_nota.produto,
                    deposito=deposito
                ).aggregate(s=Sum('quantidade'))['s'] or 0

                if saldo < item_nota.quantidade:
                    raise Exception(f'Saldo insuficiente para o produto {item_nota.produto.codigo} no depósito {deposito.nome}')

                ItemMovimentacaoEstoque.objects.create(
                    movimentacao=movimentacao,
                    produto=item_nota.produto,
                    quantidade=item_nota.quantidade,
                    preco_unitario=item_nota.preco_unitario
                )

                # Atualizar saldo
                posicao = PosicaoEstoque.objects.get(
                    produto=item_nota.produto,
                    deposito=deposito
                )
                posicao.quantidade -= item_nota.quantidade
                posicao.save()

            # Atualizar status
            nota.status = 'entregue'
            nota.save()

        return JsonResponse({
            'success': True, 
            'message': 'Entrega confirmada e estoque atualizado!'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# -----------------------------------------------------------------------------
# RELATÓRIO DE VENDAS
# -----------------------------------------------------------------------------

@login_required
def relatorio_vendas(request):
    """Relatório de vendas"""
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    cliente_id = request.GET.get('cliente_id')
    vendedor_id = request.GET.get('vendedor_id')

    queryset = NotaFiscalSaida.objects.filter(status__in=['emitida', 'entregue'])

    if data_inicio:
        queryset = queryset.filter(data_emissao__gte=data_inicio)
    if data_fim:
        queryset = queryset.filter(data_emissao__lte=data_fim)
    if cliente_id:
        queryset = queryset.filter(cliente_id=cliente_id)
    if vendedor_id:
        queryset = queryset.filter(pedido__vendedor_id=vendedor_id)

    total_vendas = queryset.aggregate(total=Sum('valor_total'))['total'] or 0

    # Agrupar por cliente
    por_cliente = queryset.values('cliente__nome').annotate(
        total=Sum('valor_total'),
        quantidade=Count('id')
    ).order_by('-total')

    # Agrupar por vendedor
    por_vendedor = queryset.values('pedido__vendedor__nome').annotate(
        total=Sum('valor_total'),
        quantidade=Count('id')
    ).order_by('-total')

    # Agrupar por mês
    por_mes = queryset.extra(
        select={'mes': "TO_CHAR(data_emissao, 'YYYY-MM')"}
    ).values('mes').annotate(
        total=Sum('valor_total'),
        quantidade=Count('id')
    ).order_by('mes')

    return render(request, 'vendas/relatorio_vendas.html', {
        'notas': queryset,
        'total_vendas': total_vendas,
        'por_cliente': por_cliente,
        'por_vendedor': por_vendedor,
        'por_mes': por_mes,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'cliente_id': cliente_id,
        'vendedor_id': vendedor_id,
        'clientes': Cliente.objects.filter(ativo=True),
        'vendedores': Vendedor.objects.filter(ativo=True)
    })


# =============================================================================
# 4. FINANCEIRO
# =============================================================================

# -----------------------------------------------------------------------------
# RELATÓRIO FINANCEIRO
# -----------------------------------------------------------------------------

@login_required
def relatorio_financeiro(request):
    """Relatório financeiro consolidado"""
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')

    hoje = timezone.now().date()
    if not data_inicio:
        data_inicio = (hoje - timedelta(days=30)).isoformat()
    if not data_fim:
        data_fim = hoje.isoformat()

    # Contas a pagar
    contas_pagar = ContaPagar.objects.filter(
        data_vencimento__gte=data_inicio,
        data_vencimento__lte=data_fim
    )
    total_pagar = contas_pagar.aggregate(total=Sum('valor'))['total'] or 0

    # Contas a receber
    contas_receber = ContaReceber.objects.filter(
        data_vencimento__gte=data_inicio,
        data_vencimento__lte=data_fim
    )
    total_receber = contas_receber.aggregate(total=Sum('valor'))['total'] or 0

    # Fluxo de caixa
    fluxo = FluxoCaixa.objects.filter(
        data__gte=data_inicio,
        data__lte=data_fim
    ).aggregate(
        entrada=Sum('valor_entrada'),
        saida=Sum('valor_saida')
    )

    context = {
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'total_pagar': total_pagar,
        'total_receber': total_receber,
        'saldo_previsto': total_receber - total_pagar,
        'contas_pagar': contas_pagar,
        'contas_receber': contas_receber,
        'fluxo_entrada': fluxo['entrada'] or 0,
        'fluxo_saida': fluxo['saida'] or 0,
    }
    return render(request, 'financeiro/relatorio_financeiro.html', context)


# -----------------------------------------------------------------------------
# FLUXO DE CAIXA
# -----------------------------------------------------------------------------

@login_required
def fluxo_caixa_manager(request):
    """Lista de lançamentos de fluxo de caixa"""
    queryset = FluxoCaixa.objects.all().order_by('-data', '-id')

    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    tipo = request.GET.get('tipo')

    if data_inicio:
        queryset = queryset.filter(data__gte=data_inicio)
    if data_fim:
        queryset = queryset.filter(data__lte=data_fim)
    if tipo:
        queryset = queryset.filter(tipo=tipo)

    paginator = Paginator(queryset, 25)
    page = request.GET.get('page', 1)
    lancamentos = paginator.get_page(page)

    # Totais
    totais = queryset.aggregate(
        entrada=Sum('valor_entrada'),
        saida=Sum('valor_saida')
    )

    return render(request, 'financeiro/fluxo_caixa_manager.html', {
        'lancamentos': lancamentos,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'tipo': tipo,
        'total_entrada': totais['entrada'] or 0,
        'total_saida': totais['saida'] or 0,
        'saldo': (totais['entrada'] or 0) - (totais['saida'] or 0)
    })


@login_required
def fluxo_caixa_form(request, pk=None):
    """Formulário de fluxo de caixa"""
    lancamento = get_object_or_404(FluxoCaixa, pk=pk) if pk else None

    if request.method == 'POST':
        form = FluxoCaixaForm(request.POST, instance=lancamento)
        if form.is_valid():
            form.save()
            messages.success(request, 'Lançamento salvo com sucesso!')
            return redirect('ERP_ServicesBI:fluxo_caixa_manager')
    else:
        form = FluxoCaixaForm(instance=lancamento)
        if not lancamento:
            form.initial['data'] = timezone.now().date()

    return render(request, 'financeiro/fluxo_caixa_form.html', {
        'form': form,
        'lancamento': lancamento,
        'is_edit': bool(pk)
    })


@login_required
def fluxo_caixa_delete(request, pk):
    """Excluir lançamento de fluxo de caixa"""
    lancamento = get_object_or_404(FluxoCaixa, pk=pk)

    if request.method == 'POST':
        lancamento.delete()
        messages.success(request, 'Lançamento excluído com sucesso!')
        return redirect('ERP_ServicesBI:fluxo_caixa_manager')

    return render(request, 'financeiro/fluxo_caixa_confirm_delete.html', {'lancamento': lancamento})


# -----------------------------------------------------------------------------
# CONTAS A PAGAR
# -----------------------------------------------------------------------------

@login_required
def conta_pagar_list(request):
    """Lista de contas a pagar"""
    queryset = ContaPagar.objects.all().order_by('data_vencimento')

    search = request.GET.get('search', '')
    status = request.GET.get('status', '')
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')

    if search:
        queryset = queryset.filter(
            Q(descricao__icontains=search) |
            Q(fornecedor__nome__icontains=search)
        )
    if status:
        queryset = queryset.filter(status=status)
    if data_inicio:
        queryset = queryset.filter(data_vencimento__gte=data_inicio)
    if data_fim:
        queryset = queryset.filter(data_vencimento__lte=data_fim)

    paginator = Paginator(queryset, 25)
    page = request.GET.get('page', 1)
    contas = paginator.get_page(page)

    # Totais por status
    totais = ContaPagar.objects.values('status').annotate(total=Sum('valor'))

    return render(request, 'financeiro/conta_pagar_manager.html', {
        'contas': contas,
        'search': search,
        'status': status,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'totais': {t['status']: t['total'] for t in totais}
    })


@login_required
def conta_pagar_add(request):
    """Adicionar conta a pagar"""
    if request.method == 'POST':
        form = ContaPagarForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Conta a pagar cadastrada com sucesso!')
            return redirect('conta_pagar_list')
    else:
        form = ContaPagarForm()
        form.initial['data_emissao'] = timezone.now().date()

    return render(request, 'financeiro/conta_pagar_form.html', {
        'form': form,
        'conta': None,
        'is_edit': False
    })


@login_required
def conta_pagar_edit(request, pk):
    """Editar conta a pagar"""
    conta = get_object_or_404(ContaPagar, pk=pk)

    if request.method == 'POST':
        form = ContaPagarForm(request.POST, instance=conta)
        if form.is_valid():
            form.save()
            messages.success(request, 'Conta a pagar atualizada com sucesso!')
            return redirect('conta_pagar_list')
    else:
        form = ContaPagarForm(instance=conta)

    return render(request, 'financeiro/conta_pagar_form.html', {
        'form': form,
        'conta': conta,
        'is_edit': True
    })


@login_required
def conta_pagar_delete(request, pk):
    """Excluir conta a pagar"""
    conta = get_object_or_404(ContaPagar, pk=pk)

    if request.method == 'POST':
        conta.delete()
        messages.success(request, 'Conta a pagar excluída com sucesso!')
        return redirect('conta_pagar_list')

    return render(request, 'financeiro/conta_pagar_confirm_delete.html', {'conta': conta})


@login_required
@require_POST
def conta_pagar_baixar(request, pk):
    """Baixar conta a pagar (marcar como paga)"""
    try:
        conta = get_object_or_404(ContaPagar, pk=pk)
        data_pagamento = request.POST.get('data_pagamento')
        valor_pago = request.POST.get('valor_pago')

        conta.status = 'pago'
        conta.data_pagamento = data_pagamento
        conta.valor_pago = valor_pago
        conta.save()

        # Criar lançamento no fluxo de caixa
        FluxoCaixa.objects.create(
            data=data_pagamento,
            tipo='saida',
            descricao=f'Pagamento: {conta.descricao}',
            valor_saida=valor_pago,
            categoria='despesa',
            conta_pagar=conta
        )

        return JsonResponse({'success': True, 'message': 'Conta baixada com sucesso!'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# -----------------------------------------------------------------------------
# CONTAS A RECEBER
# -----------------------------------------------------------------------------

@login_required
def conta_receber_list(request):
    """Lista de contas a receber"""
    queryset = ContaReceber.objects.all().order_by('data_vencimento')

    search = request.GET.get('search', '')
    status = request.GET.get('status', '')
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')

    if search:
        queryset = queryset.filter(
            Q(descricao__icontains=search) |
            Q(cliente__nome__icontains=search)
        )
    if status:
        queryset = queryset.filter(status=status)
    if data_inicio:
        queryset = queryset.filter(data_vencimento__gte=data_inicio)
    if data_fim:
        queryset = queryset.filter(data_vencimento__lte=data_fim)

    paginator = Paginator(queryset, 25)
    page = request.GET.get('page', 1)
    contas = paginator.get_page(page)

    # Totais por status
    totais = ContaReceber.objects.values('status').annotate(total=Sum('valor'))

    return render(request, 'financeiro/conta_receber_manager.html', {
        'contas': contas,
        'search': search,
        'status': status,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'totais': {t['status']: t['total'] for t in totais}
    })


@login_required
def conta_receber_add(request):
    """Adicionar conta a receber"""
    if request.method == 'POST':
        form = ContaReceberForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Conta a receber cadastrada com sucesso!')
            return redirect('conta_receber_list')
    else:
        form = ContaReceberForm()
        form.initial['data_emissao'] = timezone.now().date()

    return render(request, 'financeiro/conta_receber_form.html', {
        'form': form,
        'conta': None,
        'is_edit': False
    })


@login_required
def conta_receber_edit(request, pk):
    """Editar conta a receber"""
    conta = get_object_or_404(ContaReceber, pk=pk)

    if request.method == 'POST':
        form = ContaReceberForm(request.POST, instance=conta)
        if form.is_valid():
            form.save()
            messages.success(request, 'Conta a receber atualizada com sucesso!')
            return redirect('conta_receber_list')
    else:
        form = ContaReceberForm(instance=conta)

    return render(request, 'financeiro/conta_receber_form.html', {
        'form': form,
        'conta': conta,
        'is_edit': True
    })


@login_required
def conta_receber_delete(request, pk):
    """Excluir conta a receber"""
    conta = get_object_or_404(ContaReceber, pk=pk)

    if request.method == 'POST':
        conta.delete()
        messages.success(request, 'Conta a receber excluída com sucesso!')
        return redirect('conta_receber_list')

    return render(request, 'financeiro/conta_receber_confirm_delete.html', {'conta': conta})


@login_required
@require_POST
def conta_receber_baixar(request, pk):
    """Baixar conta a receber (marcar como recebida)"""
    try:
        conta = get_object_or_404(ContaReceber, pk=pk)
        data_recebimento = request.POST.get('data_recebimento')
        valor_recebido = request.POST.get('valor_recebido')

        conta.status = 'recebido'
        conta.data_recebimento = data_recebimento
        conta.valor_recebido = valor_recebido
        conta.save()

        # Criar lançamento no fluxo de caixa
        FluxoCaixa.objects.create(
            data=data_recebimento,
            tipo='entrada',
            descricao=f'Recebimento: {conta.descricao}',
            valor_entrada=valor_recebido,
            categoria='receita',
            conta_receber=conta
        )

        return JsonResponse({'success': True, 'message': 'Conta baixada com sucesso!'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# -----------------------------------------------------------------------------
# CONCILIAÇÃO BANCÁRIA
# -----------------------------------------------------------------------------

@login_required
def conciliacao_bancaria_manager(request):
    """Lista de conciliações bancárias"""
    queryset = ConciliacaoBancaria.objects.all().order_by('-data', '-id')

    conta_id = request.GET.get('conta_id')
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    status = request.GET.get('status')

    if conta_id:
        queryset = queryset.filter(conta_bancaria_id=conta_id)
    if data_inicio:
        queryset = queryset.filter(data__gte=data_inicio)
    if data_fim:
        queryset = queryset.filter(data__lte=data_fim)
    if status:
        queryset = queryset.filter(status=status)

    paginator = Paginator(queryset, 25)
    page = request.GET.get('page', 1)
    conciliacoes = paginator.get_page(page)

    return render(request, 'financeiro/conciliacao_bancaria_manager.html', {
        'conciliacoes': conciliacoes,
        'conta_id': conta_id,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'status': status,
        'contas': ContaBancaria.objects.filter(ativo=True)
    })


@login_required
def conciliacao_bancaria_form(request, pk=None):
    """Formulário de conciliação bancária"""
    conciliacao = get_object_or_404(ConciliacaoBancaria, pk=pk) if pk else None

    if request.method == 'POST':
        form = ConciliacaoBancariaForm(request.POST, instance=conciliacao)
        if form.is_valid():
            form.save()
            messages.success(request, 'Conciliação salva com sucesso!')
            return redirect('ERP_ServicesBI:conciliacao_bancaria_manager')
    else:
        form = ConciliacaoBancariaForm(instance=conciliacao)
        if not conciliacao:
            form.initial['data'] = timezone.now().date()

    return render(request, 'financeiro/conciliacao_bancaria_form.html', {
        'form': form,
        'conciliacao': conciliacao,
        'is_edit': bool(pk),
        'contas': ContaBancaria.objects.filter(ativo=True)
    })


@login_required
def conciliacao_bancaria_delete(request, pk):
    """Excluir conciliação bancária"""
    conciliacao = get_object_or_404(ConciliacaoBancaria, pk=pk)

    if request.method == 'POST':
        conciliacao.delete()
        messages.success(request, 'Conciliação excluída com sucesso!')
        return redirect('ERP_ServicesBI:conciliacao_bancaria_manager')

    return render(request, 'financeiro/conciliacao_bancaria_confirm_delete.html', {'conciliacao': conciliacao})


# -----------------------------------------------------------------------------
# DRE (DEMONSTRAÇÃO DO RESULTADO DO EXERCÍCIO)
# -----------------------------------------------------------------------------

@login_required
def dre_manager(request):
    """Lista de configurações DRE"""
    configuracoes = ConfiguracaoDRE.objects.all().order_by('-id')
    return render(request, 'financeiro/dre_manager.html', {'configuracoes': configuracoes})


def get_configuracao_dre_form_class():
    """Retorna a classe do formulário ConfiguracaoDREForm dinamicamente para evitar import circular"""
    from .forms import ConfiguracaoDREForm
    return ConfiguracaoDREForm


@login_required
def dre_add(request):
    """Adicionar configuração DRE"""
    ConfiguracaoDREForm = get_configuracao_dre_form_class()

    if request.method == 'POST':
        form = ConfiguracaoDREForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Configuração DRE salva com sucesso!')
            return redirect('ERP_ServicesBI:dre_manager')
    else:
        form = ConfiguracaoDREForm()

    return render(request, 'financeiro/dre_form.html', {
        'form': form,
        'configuracao': None,
        'is_edit': False
    })


@login_required
def dre_edit(request, pk):
    """Editar configuração DRE"""
    configuracao = get_object_or_404(ConfiguracaoDRE, pk=pk)
    ConfiguracaoDREForm = get_configuracao_dre_form_class()

    if request.method == 'POST':
        form = ConfiguracaoDREForm(request.POST, instance=configuracao)
        if form.is_valid():
            form.save()
            messages.success(request, 'Configuração DRE atualizada com sucesso!')
            return redirect('ERP_ServicesBI:dre_manager')
    else:
        form = ConfiguracaoDREForm(instance=configuracao)

    return render(request, 'financeiro/dre_form.html', {
        'form': form,
        'configuracao': configuracao,
        'is_edit': True
    })


@login_required
def dre_delete(request, pk):
    """Excluir configuração DRE"""
    configuracao = get_object_or_404(ConfiguracaoDRE, pk=pk)

    if request.method == 'POST':
        configuracao.delete()
        messages.success(request, 'Configuração DRE excluída com sucesso!')
        return redirect('ERP_ServicesBI:dre_manager')

    return render(request, 'financeiro/dre_confirm_delete.html', {'configuracao': configuracao})


@login_required
def dre_visualizar(request, pk):
    """Visualizar DRE calculada"""
    configuracao = get_object_or_404(ConfiguracaoDRE, pk=pk)

    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')

    if not data_inicio:
        data_inicio = (timezone.now().replace(day=1) - timedelta(days=1)).replace(day=1).date()
    if not data_fim:
        data_fim = timezone.now().date()

    # Calcular DRE baseada nos lançamentos
    lancamentos = LancamentoDRE.objects.filter(
        configuracao=configuracao,
        data__gte=data_inicio,
        data__lte=data_fim
    )

    # Agrupar por categoria
    receitas = lancamentos.filter(tipo='receita').aggregate(total=Sum('valor'))['total'] or 0
    custos = lancamentos.filter(tipo='custo').aggregate(total=Sum('valor'))['total'] or 0
    despesas = lancamentos.filter(tipo='despesa').aggregate(total=Sum('valor'))['total'] or 0

    lucro_bruto = receitas - custos
    lucro_liquido = lucro_bruto - despesas

    return render(request, 'financeiro/dre_visualizar.html', {
        'configuracao': configuracao,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'receitas': receitas,
        'custos': custos,
        'despesas': despesas,
        'lucro_bruto': lucro_bruto,
        'lucro_liquido': lucro_liquido,
        'lancamentos': lancamentos
    })


# -----------------------------------------------------------------------------
# PLANEJADO X REALIZADO
# -----------------------------------------------------------------------------

@login_required
def planejado_x_realizado_manager(request):
    """Lista de planejamentos"""
    queryset = PlanejadoRealizado.objects.all().order_by('-ano', '-mes')

    ano = request.GET.get('ano')
    mes = request.GET.get('mes')

    if ano:
        queryset = queryset.filter(ano=ano)
    if mes:
        queryset = queryset.filter(mes=mes)

    paginator = Paginator(queryset, 25)
    page = request.GET.get('page', 1)
    planejamentos = paginator.get_page(page)

    return render(request, 'financeiro/planejado_x_realizado_manager.html', {
        'planejamentos': planejamentos,
        'ano': ano,
        'mes': mes
    })


@login_required
def planejado_x_realizado_form(request, pk=None):
    """Formulário de planejado x realizado"""
    planejamento = get_object_or_404(PlanejadoRealizado, pk=pk) if pk else None

    if request.method == 'POST':
        form = PlanejadoRealizadoForm(request.POST, instance=planejamento)
        if form.is_valid():
            form.save()
            messages.success(request, 'Planejamento salvo com sucesso!')
            return redirect('ERP_ServicesBI:planejado_x_realizado_manager')
    else:
        form = PlanejadoRealizadoForm(instance=planejamento)
        if not planejamento:
            hoje = timezone.now()
            form.initial['ano'] = hoje.year
            form.initial['mes'] = hoje.month

    return render(request, 'financeiro/planejado_x_realizado_form.html', {
        'form': form,
        'planejamento': planejamento,
        'is_edit': bool(pk)
    })


@login_required
def planejado_x_realizado_delete(request, pk):
    """Excluir planejamento"""
    planejamento = get_object_or_404(PlanejadoRealizado, pk=pk)

    if request.method == 'POST':
        planejamento.delete()
        messages.success(request, 'Planejamento excluído com sucesso!')
        return redirect('ERP_ServicesBI:planejado_x_realizado_manager')

    return render(request, 'financeiro/planejado_x_realizado_confirm_delete.html', {'planejamento': planejamento})


# -----------------------------------------------------------------------------
# CATEGORIA FINANCEIRA
# -----------------------------------------------------------------------------

@login_required
def categoria_financeira_list(request):
    """Lista de categorias financeiras - PRIORIDADE 1 (trava sistema se não existir)"""
    categorias = CategoriaFinanceira.objects.all().order_by('tipo', 'nome')

    tipo = request.GET.get('tipo')
    if tipo:
        categorias = categorias.filter(tipo=tipo)

    return render(request, 'financeiro/categoria_financeira_manager.html', {
        'categorias': categorias,
        'tipo': tipo
    })


@login_required
def categoria_financeira_add(request):
    """Adicionar categoria financeira"""
    if request.method == 'POST':
        form = CategoriaFinanceiraForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Categoria financeira cadastrada com sucesso!')
            return redirect('categoria_financeira_list')
    else:
        form = CategoriaFinanceiraForm()

    return render(request, 'financeiro/categoria_financeira_form.html', {
        'form': form,
        'categoria': None,
        'is_edit': False
    })


@login_required
def categoria_financeira_edit(request, pk):
    """Editar categoria financeira"""
    categoria = get_object_or_404(CategoriaFinanceira, pk=pk)

    if request.method == 'POST':
        form = CategoriaFinanceiraForm(request.POST, instance=categoria)
        if form.is_valid():
            form.save()
            messages.success(request, 'Categoria financeira atualizada com sucesso!')
            return redirect('categoria_financeira_list')
    else:
        form = CategoriaFinanceiraForm(instance=categoria)

    return render(request, 'financeiro/categoria_financeira_form.html', {
        'form': form,
        'categoria': categoria,
        'is_edit': True
    })


@login_required
def categoria_financeira_delete(request, pk):
    """Excluir categoria financeira"""
    categoria = get_object_or_404(CategoriaFinanceira, pk=pk)

    if request.method == 'POST':
        categoria.delete()
        messages.success(request, 'Categoria financeira excluída com sucesso!')
        return redirect('categoria_financeira_list')

    return render(request, 'financeiro/categoria_financeira_confirm_delete.html', {'categoria': categoria})


# -----------------------------------------------------------------------------
# CENTRO DE CUSTO
# -----------------------------------------------------------------------------

@login_required
def centro_custo_list(request):
    """Lista de centros de custo"""
    centros = CentroCusto.objects.all().order_by('codigo')
    return render(request, 'financeiro/centro_custo_manager.html', {'centros': centros})


@login_required
def centro_custo_add(request):
    """Adicionar centro de custo"""
    if request.method == 'POST':
        form = CentroCustoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Centro de custo cadastrado com sucesso!')
            return redirect('centro_custo_list')
    else:
        form = CentroCustoForm()

    return render(request, 'financeiro/centro_custo_form.html', {
        'form': form,
        'centro': None,
        'is_edit': False
    })


@login_required
def centro_custo_edit(request, pk):
    """Editar centro de custo"""
    centro = get_object_or_404(CentroCusto, pk=pk)

    if request.method == 'POST':
        form = CentroCustoForm(request.POST, instance=centro)
        if form.is_valid():
            form.save()
            messages.success(request, 'Centro de custo atualizado com sucesso!')
            return redirect('centro_custo_list')
    else:
        form = CentroCustoForm(instance=centro)

    return render(request, 'financeiro/centro_custo_form.html', {
        'form': form,
        'centro': centro,
        'is_edit': True
    })


@login_required
def centro_custo_delete(request, pk):
    """Excluir centro de custo"""
    centro = get_object_or_404(CentroCusto, pk=pk)

    if request.method == 'POST':
        centro.delete()
        messages.success(request, 'Centro de custo excluído com sucesso!')
        return redirect('centro_custo_list')

    return render(request, 'financeiro/centro_custo_confirm_delete.html', {'centro': centro})


# -----------------------------------------------------------------------------
# CONTA BANCÁRIA
# -----------------------------------------------------------------------------

@login_required
def conta_bancaria_list(request):
    """Lista de contas bancárias"""
    contas = ContaBancaria.objects.all().order_by('nome')
    return render(request, 'financeiro/conta_bancaria_manager.html', {'contas': contas})


@login_required
def conta_bancaria_add(request):
    """Adicionar conta bancária"""
    if request.method == 'POST':
        form = ContaBancariaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Conta bancária cadastrada com sucesso!')
            return redirect('conta_bancaria_list')
    else:
        form = ContaBancariaForm()

    return render(request, 'financeiro/conta_bancaria_form.html', {
        'form': form,
        'conta': None,
        'is_edit': False
    })


@login_required
def conta_bancaria_edit(request, pk):
    """Editar conta bancária"""
    conta = get_object_or_404(ContaBancaria, pk=pk)

    if request.method == 'POST':
        form = ContaBancariaForm(request.POST, instance=conta)
        if form.is_valid():
            form.save()
            messages.success(request, 'Conta bancária atualizada com sucesso!')
            return redirect('conta_bancaria_list')
    else:
        form = ContaBancariaForm(instance=conta)

    return render(request, 'financeiro/conta_bancaria_form.html', {
        'form': form,
        'conta': conta,
        'is_edit': True
    })


@login_required
def conta_bancaria_delete(request, pk):
    """Excluir conta bancária"""
    conta = get_object_or_404(ContaBancaria, pk=pk)

    if request.method == 'POST':
        conta.delete()
        messages.success(request, 'Conta bancária excluída com sucesso!')
        return redirect('conta_bancaria_list')

    return render(request, 'financeiro/conta_bancaria_confirm_delete.html', {'conta': conta})


# -----------------------------------------------------------------------------
# FORMA DE PAGAMENTO
# -----------------------------------------------------------------------------

@login_required
def forma_pagamento_list(request):
    """Lista de formas de pagamento"""
    formas = FormaPagamento.objects.all().order_by('nome')
    return render(request, 'financeiro/forma_pagamento_manager.html', {'formas': formas})


@login_required
def forma_pagamento_add(request):
    """Adicionar forma de pagamento"""
    if request.method == 'POST':
        form = FormaPagamentoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Forma de pagamento cadastrada com sucesso!')
            return redirect('forma_pagamento_list')
    else:
        form = FormaPagamentoForm()

    return render(request, 'financeiro/forma_pagamento_form.html', {
        'form': form,
        'forma': None,
        'is_edit': False
    })


@login_required
def forma_pagamento_edit(request, pk):
    """Editar forma de pagamento"""
    forma = get_object_or_404(FormaPagamento, pk=pk)

    if request.method == 'POST':
        form = FormaPagamentoForm(request.POST, instance=forma)
        if form.is_valid():
            form.save()
            messages.success(request, 'Forma de pagamento atualizada com sucesso!')
            return redirect('forma_pagamento_list')
    else:
        form = FormaPagamentoForm(instance=forma)

    return render(request, 'financeiro/forma_pagamento_form.html', {
        'form': form,
        'forma': forma,
        'is_edit': True
    })


@login_required
def forma_pagamento_delete(request, pk):
    """Excluir forma de pagamento"""
    forma = get_object_or_404(FormaPagamento, pk=pk)

    if request.method == 'POST':
        forma.delete()
        messages.success(request, 'Forma de pagamento excluída com sucesso!')
        return redirect('forma_pagamento_list')

    return render(request, 'financeiro/forma_pagamento_confirm_delete.html', {'forma': forma})


# -----------------------------------------------------------------------------
# CONDIÇÃO DE PAGAMENTO
# -----------------------------------------------------------------------------

@login_required
def condicao_pagamento_list(request):
    """Lista de condições de pagamento"""
    condicoes = CondicaoPagamento.objects.all().order_by('nome')
    return render(request, 'financeiro/condicao_pagamento_manager.html', {'condicoes': condicoes})


@login_required
def condicao_pagamento_add(request):
    """Adicionar condição de pagamento"""
    if request.method == 'POST':
        form = CondicaoPagamentoForm(request.POST)
        if form.is_valid():
            condicao = form.save()
            # Processar parcelas se enviadas via JSON
            parcelas_json = request.POST.get('parcelas_json', '[]')
            try:
                parcelas = json.loads(parcelas_json)
                for p in parcelas:
                    ItemCondicaoPagamento.objects.create(
                        condicao=condicao,
                        numero_parcela=p['numero'],
                        dias=p['dias'],
                        percentual=p['percentual'],
                        forma_pagamento_id=p.get('forma_pagamento_id')
                    )
            except:
                pass
            messages.success(request, 'Condição de pagamento cadastrada com sucesso!')
            return redirect('condicao_pagamento_list')
    else:
        form = CondicaoPagamentoForm()

    return render(request, 'financeiro/condicao_pagamento_form.html', {
        'form': form,
        'condicao': None,
        'is_edit': False,
        'formas': FormaPagamento.objects.filter(ativo=True)
    })


@login_required
def condicao_pagamento_edit(request, pk):
    """Editar condição de pagamento"""
    condicao = get_object_or_404(CondicaoPagamento, pk=pk)

    if request.method == 'POST':
        form = CondicaoPagamentoForm(request.POST, instance=condicao)
        if form.is_valid():
            form.save()
            messages.success(request, 'Condição de pagamento atualizada com sucesso!')
            return redirect('condicao_pagamento_list')
    else:
        form = CondicaoPagamentoForm(instance=condicao)

    return render(request, 'financeiro/condicao_pagamento_form.html', {
        'form': form,
        'condicao': condicao,
        'is_edit': True,
        'formas': FormaPagamento.objects.filter(ativo=True),
        'parcelas': condicao.itens.all()
    })


@login_required
def condicao_pagamento_delete(request, pk):
    """Excluir condição de pagamento"""
    condicao = get_object_or_404(CondicaoPagamento, pk=pk)

    if request.method == 'POST':
        condicao.delete()
        messages.success(request, 'Condição de pagamento excluída com sucesso!')
        return redirect('condicao_pagamento_list')

    return render(request, 'financeiro/condicao_pagamento_confirm_delete.html', {'condicao': condicao})


@login_required
def condicao_pagamento_dados_api(request, pk):
    """API para buscar dados da condição de pagamento com parcelas"""
    try:
        condicao = get_object_or_404(CondicaoPagamento, pk=pk)

        parcelas = []
        for item in condicao.itens.all().order_by('numero_parcela'):
            parcelas.append({
                'numero': item.numero_parcela,
                'dias': item.dias,
                'percentual': float(item.percentual),
                'forma_pagamento_id': item.forma_pagamento_id,
                'forma_pagamento_nome': item.forma_pagamento.nome if item.forma_pagamento else ''
            })

        return JsonResponse({
            'success': True,
            'id': condicao.id,
            'nome': condicao.nome,
            'parcelas': parcelas
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# =============================================================================
# 5. ESTOQUE
# =============================================================================

# -----------------------------------------------------------------------------
# MOVIMENTAÇÕES DE ESTOQUE
# -----------------------------------------------------------------------------

@login_required
def movimentacao_estoque_manager(request):
    """Lista de movimentações de estoque"""
    queryset = MovimentacaoEstoque.objects.all().order_by('-data', '-id')

    tipo = request.GET.get('tipo')
    deposito_id = request.GET.get('deposito_id')
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')

    if tipo:
        queryset = queryset.filter(tipo=tipo)
    if deposito_id:
        queryset = queryset.filter(deposito_id=deposito_id)
    if data_inicio:
        queryset = queryset.filter(data__gte=data_inicio)
    if data_fim:
        queryset = queryset.filter(data__lte=data_fim)

    paginator = Paginator(queryset, 25)
    page = request.GET.get('page', 1)
    movimentacoes = paginator.get_page(page)

    return render(request, 'estoque/movimentacao_estoque_manager.html', {
        'movimentacoes': movimentacoes,
        'tipo': tipo,
        'deposito_id': deposito_id,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'depositos': Deposito.objects.filter(ativo=True)
    })


@login_required
def movimentacao_estoque_form(request, pk=None):
    """Formulário de movimentação de estoque"""
    movimentacao = get_object_or_404(MovimentacaoEstoque, pk=pk) if pk else None

    if request.method == 'POST':
        form = MovimentacaoEstoqueForm(request.POST, instance=movimentacao)
        if form.is_valid():
            movimentacao = form.save()
            messages.success(request, 'Movimentação salva com sucesso!')
            return redirect('ERP_ServicesBI:movimentacao_estoque_manager')
    else:
        form = MovimentacaoEstoqueForm(instance=movimentacao)
        if not movimentacao:
            form.initial['data'] = timezone.now()

    context = {
        'form': form,
        'movimentacao': movimentacao,
        'is_edit': bool(pk),
        'depositos': Deposito.objects.filter(ativo=True),
        'produtos': Produto.objects.filter(ativo=True)
    }
    return render(request, 'estoque/movimentacao_estoque_form.html', context)


@login_required
def movimentacao_estoque_delete(request, pk):
    """Excluir movimentação de estoque"""
    movimentacao = get_object_or_404(MovimentacaoEstoque, pk=pk)

    if request.method == 'POST':
        # Reverter saldos antes de excluir
        with transaction.atomic():
            for item in movimentacao.itens.all():
                posicao = PosicaoEstoque.objects.filter(
                    produto=item.produto,
                    deposito=movimentacao.deposito
                ).first()

                if posicao:
                    if movimentacao.tipo == 'entrada':
                        posicao.quantidade -= item.quantidade
                    else:
                        posicao.quantidade += item.quantidade
                    posicao.save()

            movimentacao.delete()

        messages.success(request, 'Movimentação excluída com sucesso!')
        return redirect('ERP_ServicesBI:movimentacao_estoque_manager')

    return render(request, 'estoque/movimentacao_estoque_confirm_delete.html', {'movimentacao': movimentacao})


@login_required
def movimentacao_estoque_dados_api(request, pk):
    """API para buscar dados da movimentação"""
    try:
        movimentacao = get_object_or_404(MovimentacaoEstoque, pk=pk)

        itens_list = []
        for item in movimentacao.itens.all():
            itens_list.append({
                'id': item.id,
                'produto_id': item.produto_id,
                'produto_codigo': item.produto.codigo if item.produto else '',
                'produto_descricao': item.produto.descricao if item.produto else '',
                'quantidade': float(item.quantidade),
                'preco_unitario': float(item.preco_unitario) if item.preco_unitario else 0,
            })

        return JsonResponse({
            'success': True,
            'movimentacao': {
                'id': movimentacao.id,
                'tipo': movimentacao.tipo,
                'data': movimentacao.data.isoformat() if movimentacao.data else None,
                'deposito_id': movimentacao.deposito_id,
                'deposito_nome': movimentacao.deposito.nome if movimentacao.deposito else '',
                'observacao': movimentacao.observacao or '',
            },
            'itens': itens_list
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def movimentacao_estoque_item_add_api(request):
    """API para adicionar item à movimentação"""
    try:
        data = json.loads(request.body)
        movimentacao_id = data.get('movimentacao_id')
        movimentacao = get_object_or_404(MovimentacaoEstoque, pk=movimentacao_id)

        item = ItemMovimentacaoEstoque.objects.create(
            movimentacao=movimentacao,
            produto_id=data.get('produto_id'),
            quantidade=data.get('quantidade'),
            preco_unitario=data.get('preco_unitario', 0)
        )

        # Atualizar saldo do estoque
        posicao, created = PosicaoEstoque.objects.get_or_create(
            produto=item.produto,
            deposito=movimentacao.deposito,
            defaults={'quantidade': 0}
        )

        if movimentacao.tipo == 'entrada':
            posicao.quantidade += item.quantidade
        else:
            posicao.quantidade -= item.quantidade
        posicao.save()

        return JsonResponse({
            'success': True, 
            'item_id': item.id
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def movimentacao_estoque_item_delete_api(request, item_id):
    """API para remover item da movimentação"""
    try:
        item = get_object_or_404(ItemMovimentacaoEstoque, pk=item_id)
        movimentacao = item.movimentacao

        # Reverter saldo
        posicao = PosicaoEstoque.objects.filter(
            produto=item.produto,
            deposito=movimentacao.deposito
        ).first()

        if posicao:
            if movimentacao.tipo == 'entrada':
                posicao.quantidade -= item.quantidade
            else:
                posicao.quantidade += item.quantidade
            posicao.save()

        item.delete()

        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# -----------------------------------------------------------------------------
# INVENTÁRIO
# -----------------------------------------------------------------------------

@login_required
def inventario_manager(request):
    """Lista de inventários - com modo contagem via ?contagem=id"""
    contagem_id = request.GET.get('contagem')

    if contagem_id:
        # Modo contagem: mostrar formulário de contagem do inventário específico
        inventario = get_object_or_404(Inventario, pk=contagem_id)
        return render(request, 'estoque/inventario_form.html', {
            'inventario': inventario,
            'itens': inventario.itens.all(),
            'modo_contagem': True
        })

    # Modo lista normal
    queryset = Inventario.objects.all().order_by('-data', '-id')

    deposito_id = request.GET.get('deposito_id')
    status = request.GET.get('status')

    if deposito_id:
        queryset = queryset.filter(deposito_id=deposito_id)
    if status:
        queryset = queryset.filter(status=status)

    paginator = Paginator(queryset, 25)
    page = request.GET.get('page', 1)
    inventarios = paginator.get_page(page)

    return render(request, 'estoque/inventario_manager.html', {
        'inventarios': inventarios,
        'deposito_id': deposito_id,
        'status': status,
        'depositos': Deposito.objects.filter(ativo=True)
    })


@login_required
def inventario_form(request, pk=None):
    """Formulário de inventário"""
    inventario = get_object_or_404(Inventario, pk=pk) if pk else None

    if request.method == 'POST':
        form = InventarioForm(request.POST, instance=inventario)
        if form.is_valid():
            inventario = form.save()

            # Se novo inventário, gerar itens automaticamente
            if not pk:
                inventario.gerar_itens_inventario()
                messages.success(request, f'Inventário criado com {inventario.itens.count()} itens!')
            else:
                messages.success(request, 'Inventário atualizado com sucesso!')

            return redirect('ERP_ServicesBI:inventario_manager')
    else:
        form = InventarioForm(instance=inventario)
        if not inventario:
            form.initial['data'] = timezone.now().date()

    return render(request, 'estoque/inventario_form.html', {
        'form': form,
        'inventario': inventario,
        'is_edit': bool(pk),
        'depositos': Deposito.objects.filter(ativo=True)
    })


@login_required
def inventario_delete(request, pk):
    """Excluir inventário"""
    inventario = get_object_or_404(Inventario, pk=pk)

    if request.method == 'POST':
        inventario.delete()
        messages.success(request, 'Inventário excluído com sucesso!')
        return redirect('ERP_ServicesBI:inventario_manager')

    return render(request, 'estoque/inventario_confirm_delete.html', {'inventario': inventario})


@login_required
@require_POST
def inventario_item_atualizar_api(request, item_id):
    """API para atualizar quantidade contada de um item"""
    try:
        item = get_object_or_404(ItemInventario, pk=item_id)
        data = json.loads(request.body)

        item.quantidade_contada = data.get('quantidade_contada', 0)
        item.observacao = data.get('observacao', '')
        item.usuario_contagem = request.user
        item.data_contagem = timezone.now()
        item.save()

        # Calcular diferença
        diferenca = item.quantidade_contada - item.quantidade_sistema

        return JsonResponse({
            'success': True,
            'diferenca': float(diferenca)
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def inventario_finalizar(request, pk):
    """Finalizar inventário e gerar ajustes de estoque"""
    try:
        inventario = get_object_or_404(Inventario, pk=pk)

        with transaction.atomic():
            # Criar movimentação de ajuste para cada item com diferença
            for item in inventario.itens.all():
                diferenca = item.quantidade_contada - item.quantidade_sistema

                if diferenca != 0:
                    tipo = 'entrada' if diferenca > 0 else 'saida'

                    movimentacao = MovimentacaoEstoque.objects.create(
                        tipo=tipo,
                        data=timezone.now(),
                        deposito=inventario.deposito,
                        observacao=f'Ajuste via Inventário {inventario.id} - Item {item.produto.codigo}'
                    )

                    ItemMovimentacaoEstoque.objects.create(
                        movimentacao=movimentacao,
                        produto=item.produto,
                        quantidade=abs(diferenca),
                        preco_unitario=item.produto.preco_custo or 0
                    )

                    # Atualizar posição
                    posicao, created = PosicaoEstoque.objects.get_or_create(
                        produto=item.produto,
                        deposito=inventario.deposito,
                        defaults={'quantidade': 0}
                    )
                    posicao.quantidade = item.quantidade_contada
                    posicao.save()

            inventario.status = 'finalizado'
            inventario.data_fim = timezone.now()
            inventario.save()

        return JsonResponse({
            'success': True,
            'message': 'Inventário finalizado e ajustes aplicados!'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# -----------------------------------------------------------------------------
# TRANSFERÊNCIAS DE ESTOQUE
# -----------------------------------------------------------------------------

@login_required
def transferencia_estoque_manager(request):
    """Lista de transferências de estoque"""
    queryset = TransferenciaEstoque.objects.all().order_by('-data', '-id')

    deposito_origem_id = request.GET.get('deposito_origem_id')
    deposito_destino_id = request.GET.get('deposito_destino_id')
    status = request.GET.get('status')

    if deposito_origem_id:
        queryset = queryset.filter(deposito_origem_id=deposito_origem_id)
    if deposito_destino_id:
        queryset = queryset.filter(deposito_destino_id=deposito_destino_id)
    if status:
        queryset = queryset.filter(status=status)

    paginator = Paginator(queryset, 25)
    page = request.GET.get('page', 1)
    transferencias = paginator.get_page(page)

    return render(request, 'estoque/transferencia_estoque_manager.html', {
        'transferencias': transferencias,
        'deposito_origem_id': deposito_origem_id,
        'deposito_destino_id': deposito_destino_id,
        'status': status,
        'depositos': Deposito.objects.filter(ativo=True)
    })


@login_required
def transferencia_estoque_form(request, pk=None):
    """Formulário de transferência de estoque"""
    transferencia = get_object_or_404(TransferenciaEstoque, pk=pk) if pk else None

    if request.method == 'POST':
        form = TransferenciaEstoqueForm(request.POST, instance=transferencia)
        if form.is_valid():
            transferencia = form.save()
            messages.success(request, 'Transferência salva com sucesso!')
            return redirect('ERP_ServicesBI:transferencia_estoque_manager')
    else:
        form = TransferenciaEstoqueForm(instance=transferencia)
        if not transferencia:
            form.initial['data'] = timezone.now().date()

    return render(request, 'estoque/transferencia_estoque_form.html', {
        'form': form,
        'transferencia': transferencia,
        'is_edit': bool(pk),
        'depositos': Deposito.objects.filter(ativo=True),
        'produtos': Produto.objects.filter(ativo=True)
    })


@login_required
def transferencia_estoque_delete(request, pk):
    """Excluir transferência de estoque"""
    transferencia = get_object_or_404(TransferenciaEstoque, pk=pk)

    if request.method == 'POST':
        # Reverter movimentações se já efetivada
        if transferencia.status == 'efetivada':
            with transaction.atomic():
                for item in transferencia.itens.all():
                    # Devolver ao origem
                    posicao_origem = PosicaoEstoque.objects.get(
                        produto=item.produto,
                        deposito=transferencia.deposito_origem
                    )
                    posicao_origem.quantidade += item.quantidade
                    posicao_origem.save()

                    # Retirar do destino
                    posicao_destino = PosicaoEstoque.objects.get(
                        produto=item.produto,
                        deposito=transferencia.deposito_destino
                    )
                    posicao_destino.quantidade -= item.quantidade
                    posicao_destino.save()

        transferencia.delete()
        messages.success(request, 'Transferência excluída com sucesso!')
        return redirect('ERP_ServicesBI:transferencia_estoque_manager')

    return render(request, 'estoque/transferencia_estoque_confirm_delete.html', {'transferencia': transferencia})


@login_required
@require_POST
def transferencia_estoque_efetivar(request, pk):
    """Efetivar transferência de estoque"""
    try:
        transferencia = get_object_or_404(TransferenciaEstoque, pk=pk)

        with transaction.atomic():
            for item in transferencia.itens.all():
                # Verificar saldo no origem
                posicao_origem = PosicaoEstoque.objects.filter(
                    produto=item.produto,
                    deposito=transferencia.deposito_origem
                ).first()

                if not posicao_origem or posicao_origem.quantidade < item.quantidade:
                    raise Exception(f'Saldo insuficiente para {item.produto.codigo} no depósito origem')

                # Retirar do origem
                posicao_origem.quantidade -= item.quantidade
                posicao_origem.save()

                # Adicionar ao destino
                posicao_destino, created = PosicaoEstoque.objects.get_or_create(
                    produto=item.produto,
                    deposito=transferencia.deposito_destino,
                    defaults={'quantidade': 0}
                )
                posicao_destino.quantidade += item.quantidade
                posicao_destino.save()

            transferencia.status = 'efetivada'
            transferencia.data_efetivacao = timezone.now()
            transferencia.save()

        return JsonResponse({
            'success': True,
            'message': 'Transferência efetivada com sucesso!'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# -----------------------------------------------------------------------------
# DEPÓSITOS
# -----------------------------------------------------------------------------

@login_required
def deposito_list(request):
    """Lista de depósitos"""
    depositos = Deposito.objects.all().order_by('nome')
    return render(request, 'estoque/deposito_manager.html', {'depositos': depositos})


@login_required
def deposito_add(request):
    """Adicionar depósito"""
    if request.method == 'POST':
        form = DepositoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Depósito cadastrado com sucesso!')
            return redirect('deposito_list')
    else:
        form = DepositoForm()

    return render(request, 'estoque/deposito_form.html', {
        'form': form,
        'deposito': None,
        'is_edit': False
    })


@login_required
def deposito_edit(request, pk):
    """Editar depósito"""
    deposito = get_object_or_404(Deposito, pk=pk)

    if request.method == 'POST':
        form = DepositoForm(request.POST, instance=deposito)
        if form.is_valid():
            form.save()
            messages.success(request, 'Depósito atualizado com sucesso!')
            return redirect('deposito_list')
    else:
        form = DepositoForm(instance=deposito)

    return render(request, 'estoque/deposito_form.html', {
        'form': form,
        'deposito': deposito,
        'is_edit': True
    })


@login_required
def deposito_delete(request, pk):
    """Excluir depósito"""
    deposito = get_object_or_404(Deposito, pk=pk)

    if request.method == 'POST':
        deposito.delete()
        messages.success(request, 'Depósito excluído com sucesso!')
        return redirect('deposito_list')

    return render(request, 'estoque/deposito_confirm_delete.html', {'deposito': deposito})


# -----------------------------------------------------------------------------
# RELATÓRIOS DE ESTOQUE
# -----------------------------------------------------------------------------

@login_required
def relatorio_estoque_posicao(request):
    """Relatório de posição de estoque"""
    deposito_id = request.GET.get('deposito_id')
    produto_id = request.GET.get('produto_id')

    queryset = PosicaoEstoque.objects.all()

    if deposito_id:
        queryset = queryset.filter(deposito_id=deposito_id)
    if produto_id:
        queryset = queryset.filter(produto_id=produto_id)

    # Apenas saldos positivos
    queryset = queryset.filter(quantidade__gt=0).order_by('produto__codigo')

    # Valor total do estoque
    valor_total = sum(p.quantidade * (p.produto.preco_custo or 0) for p in queryset)

    return render(request, 'estoque/relatorio_posicao.html', {
        'posicoes': queryset,
        'valor_total': valor_total,
        'deposito_id': deposito_id,
        'produto_id': produto_id,
        'depositos': Deposito.objects.filter(ativo=True),
        'produtos': Produto.objects.filter(ativo=True)
    })


@login_required
def relatorio_estoque_movimentacao(request):
    """Relatório de movimentações de estoque"""
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    tipo = request.GET.get('tipo')
    produto_id = request.GET.get('produto_id')

    queryset = MovimentacaoEstoque.objects.all().order_by('-data')

    if data_inicio:
        queryset = queryset.filter(data__gte=data_inicio)
    if data_fim:
        queryset = queryset.filter(data__lte=data_fim)
    if tipo:
        queryset = queryset.filter(tipo=tipo)
    if produto_id:
        queryset = queryset.filter(itens__produto_id=produto_id).distinct()

    # Totais por tipo
    totais = queryset.values('tipo').annotate(
        quantidade=Sum('itens__quantidade')
    )

    return render(request, 'estoque/relatorio_movimentacao.html', {
        'movimentacoes': queryset,
        'totais': {t['tipo']: t['quantidade'] for t in totais},
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'tipo': tipo,
        'produto_id': produto_id,
        'produtos': Produto.objects.filter(ativo=True)
    })


@login_required
def relatorio_estoque_inventario(request):
    """Relatório de inventários realizados"""
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    deposito_id = request.GET.get('deposito_id')

    queryset = Inventario.objects.all().order_by('-data')

    if data_inicio:
        queryset = queryset.filter(data__gte=data_inicio)
    if data_fim:
        queryset = queryset.filter(data__lte=data_fim)
    if deposito_id:
        queryset = queryset.filter(deposito_id=deposito_id)

    return render(request, 'estoque/relatorio_inventario.html', {
        'inventarios': queryset,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'deposito_id': deposito_id,
        'depositos': Deposito.objects.filter(ativo=True)
    })


# =============================================================================
# 6. DASHBOARD E UTILITÁRIOS
# =============================================================================

@login_required
def dashboard(request):
    """Dashboard principal do sistema"""
    hoje = timezone.now().date()

    # Vendas do mês
    vendas_mes = NotaFiscalSaida.objects.filter(
        data_emissao__year=hoje.year,
        data_emissao__month=hoje.month,
        status__in=['emitida', 'entregue']
    ).aggregate(total=Sum('valor_total'))['total'] or 0

    # Compras do mês
    compras_mes = NotaFiscalEntrada.objects.filter(
        data_recebimento__year=hoje.year,  # CORREÇÃO: data_recebimento
        data_recebimento__month=hoje.month,  # CORREÇÃO: data_recebimento
        status='recebida'
    ).aggregate(total=Sum('valor_total'))['total'] or 0

    # Contas a receber hoje
    receber_hoje = ContaReceber.objects.filter(
        data_vencimento=hoje,
        status='pendente'
    ).aggregate(total=Sum('valor'))['total'] or 0

    # Contas a pagar hoje
    pagar_hoje = ContaPagar.objects.filter(
        data_vencimento=hoje,
        status='pendente'
    ).aggregate(total=Sum('valor'))['total'] or 0

    # Pedidos pendentes
    pedidos_pendentes = PedidoVenda.objects.filter(
        status__in=['aberto', 'aprovado']
    ).count()

    # Orçamentos pendentes
    orcamentos_pendentes = Orcamento.objects.filter(
        status='aberto'
    ).count()

    # Produtos com estoque baixo
    estoque_baixo = PosicaoEstoque.objects.filter(
        quantidade_atual__lte=F('produto__estoque_minimo')
    ).select_related('produto')[:10]

    # Últimas movimentações
    ultimas_movimentacoes = MovimentacaoEstoque.objects.all().order_by('-data')[:5]

    # Gráfico de vendas por mês (últimos 6 meses)
    vendas_por_mes = []
    for i in range(5, -1, -1):
        mes = hoje - timedelta(days=i*30)
        total = NotaFiscalSaida.objects.filter(
            data_emissao__year=mes.year,
            data_emissao__month=mes.month,
            status__in=['emitida', 'entregue']
        ).aggregate(total=Sum('valor_total'))['total'] or 0
        vendas_por_mes.append({
            'mes': mes.strftime('%b/%Y'),
            'total': float(total)
        })

    context = {
        'vendas_mes': vendas_mes,
        'compras_mes': compras_mes,
        'receber_hoje': receber_hoje,
        'pagar_hoje': pagar_hoje,
        'pedidos_pendentes': pedidos_pendentes,
        'orcamentos_pendentes': orcamentos_pendentes,
        'estoque_baixo': estoque_baixo,
        'ultimas_movimentacoes': ultimas_movimentacoes,
        'vendas_por_mes': vendas_por_mes,
    }
    return render(request, 'dashboard.html', context)


@login_required
def api_dashboard_dados(request):
    """API para dados do dashboard (atualização em tempo real)"""
    hoje = timezone.now().date()

    data = {
        'vendas_mes': float(NotaFiscalSaida.objects.filter(
            data_emissao__year=hoje.year,
            data_emissao__month=hoje.month,
            status__in=['emitida', 'entregue']
        ).aggregate(total=Sum('valor_total'))['total'] or 0),
        'contas_receber_pendentes': float(ContaReceber.objects.filter(
            status='pendente'
        ).aggregate(total=Sum('valor'))['total'] or 0),
        'contas_pagar_pendentes': float(ContaPagar.objects.filter(
            status='pendente'
        ).aggregate(total=Sum('valor'))['total'] or 0),
        'pedidos_pendentes': PedidoVenda.objects.filter(
            status__in=['aberto', 'aprovado']
        ).count(),
    }

    return JsonResponse(data)


# -----------------------------------------------------------------------------
# AUTENTICAÇÃO (usando auth_views do Django - NÃO IMPLEMENTAR DIRETAMENTE)
# -----------------------------------------------------------------------------
# LoginView e LogoutView são importadas de django.contrib.auth.views
# NÃO devem ser definidas aqui - usar as views padrão do Django


# -----------------------------------------------------------------------------
# API GENÉRICAS
# -----------------------------------------------------------------------------

@login_required
def api_buscar_produtos(request):
    """API para busca de produtos (autocomplete)"""
    termo = request.GET.get('q', '')

    produtos = Produto.objects.filter(
        Q(codigo__icontains=termo) | 
        Q(descricao__icontains=termo),
        ativo=True
    )[:20]

    results = []
    for p in produtos:
        results.append({
            'id': p.id,
            'codigo': p.codigo,
            'descricao': p.descricao,
            'preco_custo': float(p.preco_custo) if p.preco_custo else 0,
            'preco_venda': float(p.preco_venda) if p.preco_venda else 0,
            'unidade': p.unidade_medida.sigla if p.unidade_medida else 'UN'
        })

    return JsonResponse({'results': results})


@login_required
def api_buscar_clientes(request):
    """API para busca de clientes (autocomplete)"""
    termo = request.GET.get('q', '')

    clientes = Cliente.objects.filter(
        Q(nome__icontains=termo) | 
        Q(cnpj_cpf__icontains=termo),
        ativo=True
    )[:20]

    results = []
    for c in clientes:
        results.append({
            'id': c.id,
            'nome': c.nome,
            'cnpj_cpf': c.cnpj_cpf,
            'cidade': c.cidade,
            'estado': c.estado
        })

    return JsonResponse({'results': results})


@login_required
def api_buscar_fornecedores(request):
    """API para busca de fornecedores (autocomplete)"""
    termo = request.GET.get('q', '')

    fornecedores = Fornecedor.objects.filter(
        Q(nome__icontains=termo) | 
        Q(cnpj_cpf__icontains=termo),
        ativo=True
    )[:20]

    results = []
    for f in fornecedores:
        results.append({
            'id': f.id,
            'nome': f.nome,
            'cnpj_cpf': f.cnpj_cpf
        })

    return JsonResponse({'results': results})

# =============================================================================
# FIM DO ARQUIVO
# =============================================================================
# ============================================================================
# 1. CADASTRO - APIs Extras
# ============================================================================

@login_required
@require_POST
def cliente_excluir_api(request, pk):
    """Exclusão AJAX de cliente."""
    cliente = get_object_or_404(Cliente, pk=pk)
    nome = cliente.nome_razao_social
    try:
        cliente.delete()
        return JsonResponse({'success': True, 'message': f'Cliente "{nome}" excluído com sucesso!'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erro ao excluir: {str(e)}'}, status=500)


@login_required
@require_POST
def fornecedor_excluir_api(request, pk):
    """Exclusão AJAX de fornecedor."""
    fornecedor = get_object_or_404(Fornecedor, pk=pk)
    try:
        nome = fornecedor.nome_razao_social
        fornecedor.delete()
        return JsonResponse({'success': True, 'message': f'Fornecedor "{nome}" excluído!'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erro: {str(e)}'}, status=400)


@login_required
@require_POST
def empresa_excluir_api(request, pk):
    """Exclusão AJAX de empresa."""
    empresa = get_object_or_404(Empresa, pk=pk)
    try:
        nome = empresa.nome_fantasia or empresa.razao_social
        empresa.delete()
        return JsonResponse({'success': True, 'message': f'Empresa "{nome}" excluída!'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erro: {str(e)}'}, status=400)


@login_required
@require_POST
def produto_excluir_api(request, pk):
    """Exclusão AJAX de produto."""
    produto = get_object_or_404(Produto, pk=pk)
    try:
        descricao = produto.descricao
        produto.delete()
        return JsonResponse({'success': True, 'message': f'Produto "{descricao}" excluído!'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erro: {str(e)}'}, status=400)


@login_required
@require_POST
def vendedor_excluir_api(request, pk):
    """Exclusão AJAX de vendedor."""
    vendedor = get_object_or_404(Vendedor, pk=pk)
    nome = vendedor.nome
    try:
        if vendedor.foto:
            vendedor.foto.delete()
        vendedor.delete()
        return JsonResponse({'success': True, 'message': f'Vendedor "{nome}" excluído!'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erro: {str(e)}'}, status=500)


@login_required
@require_POST
def categoria_produto_create_ajax(request):
    """Criação AJAX de Categoria de Produto."""
    try:
        nome = request.POST.get('nome', '').strip()
        descricao = request.POST.get('descricao', '')
        if not nome:
            return JsonResponse({'success': False, 'message': 'Nome é obrigatório'}, status=400)
        categoria = CategoriaProduto.objects.create(nome=nome, descricao=descricao, ativo=True)
        return JsonResponse({
            'success': True,
            'id': categoria.id,
            'nome': categoria.nome,
            'message': f'Categoria "{categoria.nome}" criada!'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erro: {str(e)}'}, status=500)


@login_required
@require_POST
def categoria_produto_delete_ajax(request, pk):
    """Exclusão AJAX de Categoria de Produto."""
    try:
        categoria = get_object_or_404(CategoriaProduto, pk=pk)
        if Produto.objects.filter(categoria=categoria).exists():
            return JsonResponse({'success': False, 'message': 'Categoria em uso por produtos.'}, status=400)
        nome = categoria.nome
        categoria.delete()
        return JsonResponse({'success': True, 'message': f'Categoria "{nome}" excluída!'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erro: {str(e)}'}, status=500)


@login_required
@require_POST
def api_condicao_pagamento_criar(request):
    """API para criar condição de pagamento via AJAX."""
    try:
        data = json.loads(request.body)
        descricao = data.get('descricao', '').strip()
        parcelas = data.get('parcelas', 1)
        periodicidade = data.get('periodicidade', 'mensal')
        dias_primeira_parcela = data.get('dias_primeira_parcela', 0)

        if not descricao:
            return JsonResponse({'success': False, 'message': 'Descrição é obrigatória'}, status=400)
        if not parcelas or int(parcelas) < 1 or int(parcelas) > 24:
            return JsonResponse({'success': False, 'message': 'Parcelas deve ser entre 1 e 24'}, status=400)

        periodos_validos = [p[0] for p in CondicaoPagamento.PERIODICIDADE_CHOICES]
        if periodicidade not in periodos_validos:
            return JsonResponse({'success': False, 'message': 'Periodicidade inválida'}, status=400)

        condicao = CondicaoPagamento.objects.create(
            descricao=descricao,
            parcelas=int(parcelas),
            periodicidade=periodicidade,
            dias_primeira_parcela=int(dias_primeira_parcela),
            ativo=True
        )
        return JsonResponse({
            'success': True,
            'id': condicao.id,
            'descricao': condicao.descricao,
            'message': f'Condição "{condicao.descricao}" criada!'
        })
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Dados inválidos'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erro: {str(e)}'}, status=500)


@login_required
@require_POST
def api_condicao_pagamento_excluir(request, pk):
    """API para excluir condição de pagamento."""
    try:
        condicao = get_object_or_404(CondicaoPagamento, pk=pk)
        descricao = condicao.descricao
        if condicao.clientes_condicao.exists():
            return JsonResponse({'success': False, 'message': f'Condição "{descricao}" em uso por clientes.'}, status=400)
        condicao.delete()
        return JsonResponse({'success': True, 'message': f'Condição "{descricao}" excluída!'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erro: {str(e)}'}, status=500)


@login_required
@require_POST
def api_forma_pagamento_criar(request):
    """API para criar forma de pagamento via AJAX."""
    try:
        data = json.loads(request.body)
        descricao = data.get('descricao', '').strip()
        tipo = data.get('tipo', '').strip()

        if not descricao:
            return JsonResponse({'success': False, 'message': 'Descrição é obrigatória'}, status=400)
        if not tipo:
            return JsonResponse({'success': False, 'message': 'Tipo é obrigatório'}, status=400)

        tipos_validos = [t[0] for t in FormaPagamento.TIPO_CHOICES]
        if tipo not in tipos_validos:
            return JsonResponse({'success': False, 'message': 'Tipo inválido'}, status=400)

        forma = FormaPagamento.objects.create(descricao=descricao, tipo=tipo, ativo=True)
        return JsonResponse({
            'success': True,
            'id': forma.id,
            'descricao': f'{forma.descricao} ({forma.get_tipo_display()})',
            'message': f'Forma "{forma.descricao}" criada!'
        })
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Dados inválidos'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erro: {str(e)}'}, status=500)


@login_required
@require_POST
def api_forma_pagamento_excluir(request, pk):
    """API para excluir forma de pagamento."""
    try:
        forma = get_object_or_404(FormaPagamento, pk=pk)
        descricao = forma.descricao
        if forma.clientes_forma.exists():
            return JsonResponse({'success': False, 'message': f'Forma "{descricao}" em uso por clientes.'}, status=400)
        forma.delete()
        return JsonResponse({'success': True, 'message': f'Forma "{descricao}" excluída!'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erro: {str(e)}'}, status=500)


@login_required
@require_POST
def api_criar_projeto(request):
    """API para criar projeto via AJAX."""
    try:
        data = json.loads(request.body)
        nome = data.get('nome', '').strip()
        descricao = data.get('descricao', '').strip()

        if not nome:
            return JsonResponse({'success': False, 'message': 'Nome é obrigatório'}, status=400)

        projeto = Projeto.objects.create(
            nome=nome,
            descricao=descricao,
            ativo=True
        )

        return JsonResponse({
            'success': True,
            'id': projeto.id,
            'nome': projeto.nome,
            'message': f'Projeto "{projeto.nome}" criado!'
        })
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'JSON inválido'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erro: {str(e)}'}, status=500)


# ============================================================================
# 2. COMPRAS - APIs Extras
# ============================================================================

@login_required
@require_POST
def cotacao_salvar_api(request):
    """API unificada para salvar cotação."""
    try:
        if len(request.body) > 10 * 1024 * 1024:
            return JsonResponse({'success': False, 'message': 'Payload muito grande'}, status=413)

        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'JSON inválido'}, status=400)

        titulo = data.get('titulo', '').strip()
        if not titulo:
            return JsonResponse({'success': False, 'message': 'Título é obrigatório'}, status=400)

        pk = data.get('id')

        with transaction.atomic():
            if pk:
                cotacao = get_object_or_404(CotacaoMae, pk=pk)
            else:
                cotacao = CotacaoMae(solicitante=request.user)

            cotacao.titulo = titulo[:255]
            cotacao.setor = data.get('setor', '').strip()[:100]
            cotacao.observacoes = data.get('observacoes', '').strip()[:2000]

            if data.get('data_limite'):
                try:
                    cotacao.data_limite_resposta = datetime.fromisoformat(
                        data['data_limite'].replace('Z', '+00:00')
                    ).date()
                except (ValueError, AttributeError):
                    pass

            status_enviado = data.get('status', 'em_analise').lower()
            status_validos = ['em_analise', 'respondida', 'concluida', 'cancelada']
            cotacao.status = status_enviado if status_enviado in status_validos else 'em_analise'
            cotacao.save()

            return JsonResponse({
                'success': True,
                'id': cotacao.pk,
                'numero': cotacao.numero,
                'message': 'Cotação salva com sucesso!'
            })
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erro interno: {str(e)}'}, status=500)


@login_required
@require_POST
def cotacao_excluir_api(request, pk):
    """Exclusão AJAX de cotação."""
    try:
        get_object_or_404(CotacaoMae, pk=pk).delete()
        return JsonResponse({'success': True, 'message': 'Cotação excluída!'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erro ao excluir cotação: {str(e)}'}, status=500)


@login_required
@require_POST
def cotacao_concluir_api(request, pk):
    """Concluir cotação via API."""
    try:
        cotacao = get_object_or_404(CotacaoMae, pk=pk)
        cotacao.status = 'concluida'
        cotacao.save()
        return JsonResponse({'success': True, 'message': 'Cotação concluída!'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erro ao concluir cotação: {str(e)}'}, status=500)


@login_required
@require_GET
def cotacao_comparativo_api(request, pk):
    """API para retornar dados comparativos da cotação."""
    try:
        cotacao = get_object_or_404(CotacaoMae, pk=pk)

        cotacoes_fornecedor = CotacaoFornecedor.objects.filter(
            cotacao=cotacao
        ).select_related('fornecedor').prefetch_related('itens')

        comparativo_data = []
        for cot_forn in cotacoes_fornecedor:
            valor_total = sum(
                item.valor_total if item.valor_total else Decimal('0')
                for item in cot_forn.itens.all()
            )

            comparativo_data.append({
                'fornecedor_id': cot_forn.fornecedor_id,
                'fornecedor_nome': cot_forn.fornecedor.nome_razao_social,
                'fornecedor_fantasia': cot_forn.fornecedor.nome_fantasia or '',
                'valor_total': float(valor_total),
                'data_resposta': cot_forn.data_resposta.isoformat() if cot_forn.data_resposta else None,
                'status': cot_forn.status or 'pendente',
                'total_itens': cot_forn.itens.count(),
            })

        comparativo_data = sorted(comparativo_data, key=lambda x: x['valor_total'])

        return JsonResponse({
            'success': True,
            'cotacao_id': cotacao.id,
            'cotacao_numero': cotacao.numero,
            'comparativo': comparativo_data,
            'total_fornecedores': len(comparativo_data),
        })
    except CotacaoMae.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Cotação não encontrada'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erro: {str(e)}'}, status=500)


@login_required
@require_POST
@csrf_protect
def cotacao_gerar_pedidos(request, pk):
    """Gera pedidos de compra a partir dos itens selecionados da cotação."""
    try:
        cotacao = get_object_or_404(CotacaoMae, pk=pk)
        data = json.loads(request.body)
        pedidos_data = data.get('pedidos', [])

        if not pedidos_data:
            return JsonResponse({'success': False, 'message': 'Nenhum pedido para gerar'}, status=400)

        pedidos_gerados = []
        with transaction.atomic():
            for pedido_data in pedidos_data[:50]:
                fornecedor_id = pedido_data.get('fornecedor_id')
                itens_data = pedido_data.get('itens', [])

                if not fornecedor_id or not itens_data:
                    continue

                try:
                    fornecedor = Fornecedor.objects.get(id=fornecedor_id, ativo=True)
                except Fornecedor.DoesNotExist:
                    continue

                pedido = PedidoCompra.objects.create(
                    fornecedor=fornecedor,
                    cotacao_mae=cotacao,
                    data_prevista_entrega=timezone.now().date() + timedelta(days=15),
                    status='rascunho',
                    solicitante=request.user,
                )

                for item_data in itens_data[:1000]:
                    try:
                        quantidade = Decimal(str(item_data.get('quantidade', 1)))
                        preco_unitario = Decimal(str(item_data.get('preco_unitario', 0)))

                        if quantidade <= 0 or preco_unitario < 0:
                            continue

                        ItemPedidoCompra.objects.create(
                            pedido=pedido,
                            descricao=item_data.get('produto_nome', '')[:255],
                            quantidade=quantidade,
                            preco_unitario=preco_unitario,
                            preco_total=quantidade * preco_unitario,
                        )
                    except (ValueError, TypeError):
                        continue

                if hasattr(pedido, 'calcular_total'):
                    pedido.calcular_total()
                pedidos_gerados.append({'id': pedido.pk, 'numero': pedido.numero})

        if pedidos_gerados:
            cotacao.status = 'concluida'
            cotacao.save()

        return JsonResponse({
            'success': True,
            'pedidos': pedidos_gerados,
            'message': f'{len(pedidos_gerados)} pedido(s) gerado(s)!'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erro: {str(e)}'}, status=500)


@login_required
@require_POST
def nota_fiscal_excluir_api(request, pk):
    """Excluir NF de entrada via API."""
    try:
        nota = get_object_or_404(NotaFiscalEntrada, pk=pk)
        if nota.status == 'confirmada':
            return JsonResponse({'success': False, 'message': 'NF confirmada. Cancele primeiro.'}, status=400)
        nota.delete()
        return JsonResponse({'success': True, 'message': 'NF excluída!'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=400)


@login_required
@require_POST
def nota_fiscal_confirmar_api(request, pk):
    """Confirma NF e atualiza estoque."""
    try:
        nota = get_object_or_404(NotaFiscalEntrada, pk=pk)
        if nota.status == 'confirmada':
            return JsonResponse({'success': False, 'message': 'Já confirmada'}, status=400)
        if nota.itens.count() == 0:
            return JsonResponse({'success': False, 'message': 'NF sem itens'}, status=400)

        with transaction.atomic():
            nota.status = 'confirmada'
            nota.save()
            if hasattr(nota, 'atualizar_estoque'):
                nota.atualizar_estoque()

            if hasattr(nota, 'pedido_origem') and nota.pedido_origem:
                nota.pedido_origem.status = 'recebido'
                nota.pedido_origem.movimento_estoque_gerado = True
                nota.pedido_origem.nota_fiscal_vinculada = True
                nota.pedido_origem.save()

            if nota.valor_total > 0 and not ContaPagar.objects.filter(nota_fiscal=nota).exists():
                ContaPagar.objects.create(
                    descricao=f'NF {nota.numero_nf} - {nota.fornecedor.nome_razao_social}',
                    fornecedor=nota.fornecedor,
                    data_vencimento=timezone.now().date() + timedelta(days=30),
                    valor_original=nota.valor_total if hasattr(nota, 'valor_total') else 0,
                    status='pendente',
                )

        return JsonResponse({'success': True, 'message': 'NF confirmada! Estoque atualizado.'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erro: {str(e)}'}, status=400)


# ============================================================================
# 3. VENDAS - APIs Extras
# ============================================================================

@login_required
@require_POST
def orcamento_excluir_api(request, pk):
    """Excluir orçamento via API."""
    orcamento = get_object_or_404(Orcamento, pk=pk)
    try:
        numero = orcamento.numero
        orcamento.delete()
        return JsonResponse({'success': True, 'message': f'Orçamento {numero} excluído!'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@require_POST
def pedido_venda_excluir_api(request, pk):
    """Excluir pedido de venda via API."""
    pedido = get_object_or_404(PedidoVenda, pk=pk)
    try:
        numero = pedido.numero
        pedido.delete()
        return JsonResponse({'success': True, 'message': f'Pedido {numero} excluído!'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@require_POST
def nota_fiscal_saida_confirmar_api(request, pk):
    """Confirma NF de saída e atualiza estoque."""
    try:
        nota = get_object_or_404(NotaFiscalSaida, pk=pk)
        if nota.status == 'confirmada':
            return JsonResponse({'success': False, 'message': 'Já confirmada'}, status=400)
        if nota.itens.count() == 0:
            return JsonResponse({'success': False, 'message': 'NF sem itens'}, status=400)

        with transaction.atomic():
            nota.status = 'confirmada'
            nota.data_saida = timezone.now().date()
            nota.save()
            if hasattr(nota, '_confirmar_saida'):
                nota._confirmar_saida()

            if nota.pedido_venda:
                nota.pedido_venda.status = 'entregue'
                nota.pedido_venda.nota_fiscal_vinculada = True
                nota.pedido_venda.save()

            if nota.valor_total > 0 and not ContaReceber.objects.filter(nota_fiscal_saida=nota).exists():
                ContaReceber.objects.create(
                    descricao=f'NF {nota.numero_nf} - {nota.cliente.nome_razao_social}',
                    cliente=nota.cliente,
                    data_vencimento=timezone.now().date() + timedelta(days=30),
                    valor_original=nota.valor_total if hasattr(nota, 'valor_total') else 0,
                    status='pendente',
                )

        return JsonResponse({'success': True, 'message': 'NF confirmada! Estoque atualizado.'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erro: {str(e)}'}, status=400)


# ============================================================================
# 4. FINANCEIRO - Views e APIs Extras
# ============================================================================

@login_required
def fluxo_caixa_list(request):
    """Listagem do fluxo de caixa."""
    hoje = timezone.now().date()
    data_inicio = _parse_date(request.GET.get('data_inicio')) or (hoje - timedelta(days=30))
    data_fim = _parse_date(request.GET.get('data_fim')) or (hoje + timedelta(days=30))

    movimentacoes = MovimentoCaixa.objects.filter(data__range=[data_inicio, data_fim]).order_by('-data')
    contas_receber = ContaReceber.objects.filter(data_vencimento__range=[data_inicio, data_fim], status__in=['pendente', 'parcial'])
    contas_pagar = ContaPagar.objects.filter(data_vencimento__range=[data_inicio, data_fim], status__in=['pendente', 'parcial'])

    saldo_data = MovimentoCaixa.objects.filter(data__lte=hoje).values('tipo').annotate(total=Sum('valor'))
    ent = sum(m['total'] for m in saldo_data if m['tipo'] == 'entrada')
    sai = sum(m['total'] for m in saldo_data if m['tipo'] == 'saida')

    context = {
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'movimentacoes': movimentacoes,
        'contas_receber': contas_receber,
        'contas_pagar': contas_pagar,
        'saldo_atual': ent - sai,
        'total_receber': contas_receber.aggregate(total=Sum('valor_saldo'))['total'] or 0,
        'total_pagar': contas_pagar.aggregate(total=Sum('valor_saldo'))['total'] or 0,
    }
    return render(request, 'financeiro/fluxo_caixa_manager.html', context)


@login_required
def fluxo_caixa_add(request):
    """Adicionar movimentação de caixa."""
    if request.method == 'POST':
        form = MovimentoCaixaForm(request.POST)
        if form.is_valid():
            mov = form.save(commit=False)
            mov.usuario = request.user
            mov.save()
            messages.success(request, 'Movimentação registrada!')
            return redirect('ERP_ServicesBI:fluxo_caixa_list')
    else:
        form = MovimentoCaixaForm()
    return render(request, 'financeiro/fluxo_caixa_form.html', {'form': form, 'titulo': 'Nova Movimentação'})


def _parse_date(date_str):
    """Helper para converter string YYYY-MM-DD para date object."""
    if not date_str:
        return None
    try:
        from datetime import datetime
        return datetime.strptime(date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return None


@login_required
def configuracao_dre_list(request):
    """Listagem de configurações de DRE disponíveis."""
    configuracoes = ConfiguracaoDRE.objects.all().order_by('nome')

    context = {
        'configuracoes': configuracoes,
        'total': configuracoes.count(),
    }
    return render(request, 'financeiro/configuracao_dre_list.html', context)


@login_required
def configuracao_dre_form(request, pk=None):
    """Criar/editar configuração de DRE."""
    config = get_object_or_404(ConfiguracaoDRE, pk=pk) if pk else None

    if request.method == 'POST':
        try:
            nome = request.POST.get('nome', '').strip()
            descricao = request.POST.get('descricao', '').strip()

            if not nome:
                messages.error(request, 'Nome é obrigatório')
                return redirect('ERP_ServicesBI:configuracao_dre_form')

            if pk:
                config.nome = nome
                config.descricao = descricao
                config.save()
                messages.success(request, 'Configuração atualizada!')
            else:
                config = ConfiguracaoDRE.objects.create(
                    nome=nome,
                    descricao=descricao,
                    ativo=True
                )
                messages.success(request, 'Configuração criada!')

            return redirect('ERP_ServicesBI:configuracao_dre_list')
        except Exception as e:
            messages.error(request, f'Erro: {str(e)}')

    context = {
        'config': config,
        'titulo': 'Editar Configuração de DRE' if config else 'Nova Configuração de DRE',
    }
    return render(request, 'financeiro/configuracao_dre_form.html', context)


@login_required
def dre_list(request):
    """Listagem de relatórios DRE gerados."""
    from datetime import date
    periodo = request.GET.get('periodo', 'mensal')
    ano = request.GET.get('ano', str(date.today().year))
    mes = request.GET.get('mes', str(date.today().month))

    try:
        ano = int(ano)
        mes = int(mes)
        if mes < 1 or mes > 12:
            mes = date.today().month
    except (ValueError, TypeError):
        ano = date.today().year
        mes = date.today().month

    relatorios = RelatorioDRE.objects.filter(
        ano=ano, mes=mes, periodo=periodo
    ).order_by('-data_criacao')

    context = {
        'relatorios': relatorios,
        'periodo': periodo,
        'ano': ano,
        'mes': mes,
        'anos': range(ano - 5, ano + 2),
        'meses': range(1, 13),
    }
    return render(request, 'financeiro/dre_list.html', context)


@login_required
def dre_relatorio(request, pk):
    """Visualizar relatório DRE completo."""
    relatorio = get_object_or_404(RelatorioDRE, pk=pk)
    linhas = ItemRelatorioDRE.objects.filter(relatorio=relatorio).order_by('sequencia')

    total_receita = sum(l.valor_realizado for l in linhas if l.tipo == 'receita')
    total_despesa = sum(l.valor_realizado for l in linhas if l.tipo == 'despesa')
    lucro_liquido = total_receita - total_despesa

    context = {
        'relatorio': relatorio,
        'linhas': linhas,
        'total_receita': total_receita,
        'total_despesa': total_despesa,
        'lucro_liquido': lucro_liquido,
        'margem': (lucro_liquido / total_receita * 100) if total_receita > 0 else 0,
    }
    return render(request, 'financeiro/dre_relatorio.html', context)


@login_required
def dre_comparativo(request):
    """Comparativo de DREs entre períodos."""
    periodo1_id = request.GET.get('periodo1')
    periodo2_id = request.GET.get('periodo2')

    relatorio1 = get_object_or_404(RelatorioDRE, pk=periodo1_id) if periodo1_id else None
    relatorio2 = get_object_or_404(RelatorioDRE, pk=periodo2_id) if periodo2_id else None

    linhas1 = []
    linhas2 = []

    if relatorio1:
        linhas1 = ItemRelatorioDRE.objects.filter(relatorio=relatorio1).order_by('sequencia')
    if relatorio2:
        linhas2 = ItemRelatorioDRE.objects.filter(relatorio=relatorio2).order_by('sequencia')

    context = {
        'relatorio1': relatorio1,
        'relatorio2': relatorio2,
        'linhas1': linhas1,
        'linhas2': linhas2,
        'relatorios': RelatorioDRE.objects.all().order_by('-ano', '-mes')[:20],
    }
    return render(request, 'financeiro/dre_comparativo.html', context)


@login_required
def planejado_x_realizado_list(request):
    """Listagem de comparativos planejado vs realizado."""
    orcamentos = OrcamentoProjeto.objects.all().order_by('-ano', '-mes')

    paginator = Paginator(orcamentos, 25)
    page_obj = paginator.get_page(request.GET.get('page'))

    context = {
        'orcamentos': page_obj,
        'page_obj': page_obj,
    }
    return render(request, 'financeiro/planejado_x_realizado_list.html', context)


@login_required
def planejado_x_realizado_add(request):
    """Criar novo orçamento de projeto."""
    if request.method == 'POST':
        form = OrcamentoProjetoForm(request.POST)
        if form.is_valid():
            orcamento = form.save(commit=False)
            orcamento.criado_por = request.user
            orcamento.save()
            messages.success(request, 'Orçamento criado!')
            return redirect('ERP_ServicesBI:planejado_x_realizado_list')
    else:
        form = OrcamentoProjetoForm()

    context = {
        'form': form,
        'projetos': Projeto.objects.filter(ativo=True),
        'centros_custo': CentroCusto.objects.filter(ativo=True),
    }
    return render(request, 'financeiro/planejado_x_realizado_form.html', context)


@login_required
def planejado_x_realizado_edit(request, pk):
    """Editar orçamento de projeto."""
    orcamento = get_object_or_404(OrcamentoProjeto, pk=pk)

    if request.method == 'POST':
        form = OrcamentoProjetoForm(request.POST, instance=orcamento)
        if form.is_valid():
            form.save()
            messages.success(request, 'Orçamento atualizado!')
            return redirect('ERP_ServicesBI:planejado_x_realizado_list')
    else:
        form = OrcamentoProjetoForm(instance=orcamento)

    context = {
        'form': form,
        'orcamento': orcamento,
        'projetos': Projeto.objects.filter(ativo=True),
        'centros_custo': CentroCusto.objects.filter(ativo=True),
    }
    return render(request, 'financeiro/planejado_x_realizado_form.html', context)


@login_required
def planejado_x_realizado_excel(request):
    """Exportar comparativo planejado vs realizado para Excel."""
    from datetime import date
    ano = request.GET.get('ano', str(date.today().year))
    mes = request.GET.get('mes', str(date.today().month))

    try:
        ano = int(ano)
        mes = int(mes)
    except (ValueError, TypeError):
        ano = date.today().year
        mes = date.today().month

    orcamentos = OrcamentoProjeto.objects.filter(ano=ano, mes=mes)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="planejado_realizado_{ano}_{mes:02d}.xlsx"'

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Comparativo'

        headers = ['Projeto', 'Centro de Custo', 'Planejado', 'Realizado', 'Variação', 'Variação %']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')

        for row, orc in enumerate(orcamentos, 2):
            planejado = orc.valor_planejado if hasattr(orc, 'valor_planejado') else 0
            realizado = orc.valor_realizado if hasattr(orc, 'valor_realizado') else 0
            variacao = realizado - planejado
            var_percent = (variacao / planejado * 100) if planejado != 0 else 0

            ws.cell(row=row, column=1, value=getattr(orc.projeto, 'nome', 'N/A') if hasattr(orc, 'projeto') else 'N/A')
            ws.cell(row=row, column=2, value=getattr(orc.centro_custo, 'nome', 'N/A') if hasattr(orc, 'centro_custo') else 'N/A')
            ws.cell(row=row, column=3, value=planejado)
            ws.cell(row=row, column=4, value=realizado)
            ws.cell(row=row, column=5, value=variacao)
            ws.cell(row=row, column=6, value=f'{var_percent:.2f}%')

        for col in range(1, 7):
            ws.column_dimensions[chr(64 + col)].width = 15

        wb.save(response)
    except ImportError:
        response.write(b'openpyxl not installed')

    return response


@login_required
def conciliacao_bancaria_list(request):
    """Listagem de conciliações bancárias."""
    conta_id = request.GET.get('conta')
    status = request.GET.get('status', 'pendente')

    conciliacoes = ExtratoBancario.objects.all().order_by('-data_fim')

    if conta_id:
        conciliacoes = conciliacoes.filter(conta_id=conta_id)

    if status:
        conciliacoes = conciliacoes.filter(status=status)

    paginator = Paginator(conciliacoes, 25)
    page_obj = paginator.get_page(request.GET.get('page'))

    context = {
        'conciliacoes': page_obj,
        'page_obj': page_obj,
        'contas': ContaBancaria.objects.filter(ativo=True),
        'conta_id': conta_id,
        'status': status,
    }
    return render(request, 'financeiro/conciliacao_bancaria_manager.html', context)


@login_required
def conciliacao_bancaria_add(request):
    """Criar nova conciliação bancária."""
    if request.method == 'POST':
        form = ExtratoBancarioForm(request.POST)
        if form.is_valid():
            extrato = form.save(commit=False)
            extrato.usuario = request.user
            extrato.status = 'pendente'
            extrato.save()

            messages.success(request, 'Extrato registrado! Proceda com a conciliação.')
            return redirect('ERP_ServicesBI:conciliacao_bancaria_detail', pk=extrato.pk)
    else:
        form = ExtratoBancarioForm()

    context = {
        'form': form,
        'contas': ContaBancaria.objects.filter(ativo=True),
    }
    return render(request, 'financeiro/conciliacao_bancaria_form.html', context)


@login_required
def conciliacao_bancaria_detail(request, pk):
    """Detalhe e processamento de conciliação bancária."""
    extrato = get_object_or_404(ExtratoBancario, pk=pk)

    lancamentos = LancamentoExtrato.objects.filter(extrato=extrato).select_related(
        'lancamento_vinculado'
    ).order_by('-data')

    contas_pagar = ContaPagar.objects.filter(
        status__in=['pendente', 'parcial'],
        data_vencimento__lte=extrato.data_fim
    ).order_by('-data_vencimento')

    contas_receber = ContaReceber.objects.filter(
        status__in=['pendente', 'parcial'],
        data_vencimento__lte=extrato.data_fim
    ).order_by('-data_vencimento')

    saldo_extratos_anteriores = ExtratoBancario.objects.filter(
        conta=extrato.conta,
        data_fim__lt=extrato.data_inicio
    ).aggregate(saldo=Sum('saldo_final'))['saldo'] or 0

    context = {
        'extrato': extrato,
        'lancamentos': lancamentos,
        'contas_pagar': contas_pagar,
        'contas_receber': contas_receber,
        'saldo_anterior': saldo_extratos_anteriores,
        'total_lancamentos': lancamentos.aggregate(total=Sum('valor'))['total'] or 0,
    }
    return render(request, 'financeiro/conciliacao_bancaria_detail.html', context)


@login_required
@require_POST
def conciliacao_bancaria_vincular(request, pk):
    """Vincular lançamento do extrato a uma conta."""
    try:
        extrato = get_object_or_404(ExtratoBancario, pk=pk)
        data = json.loads(request.body)

        lancamento_id = data.get('lancamento_id')
        conta_pagar_id = data.get('conta_pagar_id')
        conta_receber_id = data.get('conta_receber_id')

        lancamento = get_object_or_404(LancamentoExtrato, pk=lancamento_id, extrato=extrato)

        if conta_pagar_id:
            conta = get_object_or_404(ContaPagar, pk=conta_pagar_id)
            lancamento.lancamento_vinculado_content_type = ContentType.objects.get_for_model(ContaPagar)
            lancamento.lancamento_vinculado_object_id = conta_pagar_id
            lancamento.save()
            return JsonResponse({'success': True, 'message': 'Vinculado a conta a pagar!'})

        elif conta_receber_id:
            conta = get_object_or_404(ContaReceber, pk=conta_receber_id)
            lancamento.lancamento_vinculado_content_type = ContentType.objects.get_for_model(ContaReceber)
            lancamento.lancamento_vinculado_object_id = conta_receber_id
            lancamento.save()
            return JsonResponse({'success': True, 'message': 'Vinculado a conta a receber!'})

        return JsonResponse({'success': False, 'message': 'Nenhuma conta selecionada'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erro: {str(e)}'}, status=400)


@login_required
@require_POST
def conciliacao_bancaria_processar(request, pk):
    """Processar conciliação (marcar como conciliada)."""
    try:
        extrato = get_object_or_404(ExtratoBancario, pk=pk)

        with transaction.atomic():
            lancamentos = LancamentoExtrato.objects.filter(extrato=extrato)
            vinculados = lancamentos.filter(lancamento_vinculado_content_type__isnull=False).count()
            total = lancamentos.count()

            if vinculados < total:
                return JsonResponse({
                    'success': False,
                    'message': f'Ainda faltam {total - vinculados} lançamento(s) vincular.'
                }, status=400)

            extrato.status = 'conciliada'
            extrato.data_conciliacao = timezone.now().date()
            extrato.conciliado_por = request.user
            extrato.save()

        messages.success(request, 'Conciliação finalizada com sucesso!')
        return JsonResponse({'success': True, 'message': 'Conciliação finalizada!'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erro: {str(e)}'}, status=400)


# ============================================================================
# 5. ESTOQUE - Views e APIs Extras
# ============================================================================

@login_required
def movimentacao_estoque_list(request):
    """Lista de movimentações de estoque."""
    search = request.GET.get('search', '')

    movimentacoes = MovimentacaoEstoque.objects.select_related(
        'produto', 'deposito_origem', 'deposito_destino'
    ).all().order_by('-data')

    if search:
        movimentacoes = movimentacoes.filter(
            Q(produto__descricao__icontains=search) |
            Q(motivo__icontains=search)
        )

    paginator = Paginator(movimentacoes, 25)
    page_obj = paginator.get_page(request.GET.get('page'))

    context = {
        'movimentacoes': page_obj,
        'page_obj': page_obj,
        'total': movimentacoes.count(),
        'search': search,
    }
    return render(request, 'estoque/movimentacao_estoque_manager.html', context)


@login_required
def movimentacao_estoque_add(request):
    """Criar movimentação de estoque."""
    if request.method == 'POST':
        form = MovimentacaoEstoqueForm(request.POST)
        if form.is_valid():
            movimentacao = form.save(commit=False)
            movimentacao.usuario = request.user
            movimentacao.save()

            messages.success(request, 'Movimentação registrada com sucesso!')
            return redirect('ERP_ServicesBI:movimentacao_estoque_list')
    else:
        form = MovimentacaoEstoqueForm()

    context = {
        'form': form,
        'produtos': Produto.objects.filter(ativo=True),
        'depositos': Deposito.objects.filter(ativo=True),
    }
    return render(request, 'estoque/movimentacao_estoque_form.html', context)


@login_required
def inventario_list(request):
    """Listagem de inventários."""
    inventarios = Inventario.objects.all().order_by('-data')

    paginator = Paginator(inventarios, 25)
    page_obj = paginator.get_page(request.GET.get('page'))

    context = {
        'inventarios': page_obj,
        'page_obj': page_obj,
        'total': inventarios.count(),
    }
    return render(request, 'estoque/inventario_manager.html', context)


@login_required
def inventario_add(request):
    """Criar novo inventário."""
    if request.method == 'POST':
        form = InventarioForm(request.POST)
        if form.is_valid():
            inventario = form.save(commit=False)
            inventario.usuario = request.user
            inventario.save()

            messages.success(request, 'Inventário iniciado!')
            return redirect('ERP_ServicesBI:inventario_list')
    else:
        form = InventarioForm()

    context = {
        'form': form,
        'depositos': Deposito.objects.filter(ativo=True),
    }
    return render(request, 'estoque/inventario_form.html', context)


@login_required
def transferencia_list(request):
    """Listagem de transferências."""
    transferencias = TransferenciaEstoque.objects.select_related('usuario').prefetch_related('itens').all().order_by('-data')

    paginator = Paginator(transferencias, 25)
    page_obj = paginator.get_page(request.GET.get('page'))

    context = {
        'transferencias': page_obj,
        'page_obj': page_obj,
        'total': transferencias.count(),
    }
    return render(request, 'estoque/transferencia_manager.html', context)


@login_required
def transferencia_add(request):
    """Criar nova transferência."""
    if request.method == 'POST':
        form = TransferenciaEstoqueForm(request.POST)
        if form.is_valid():
            transferencia = form.save(commit=False)
            transferencia.usuario = request.user
            transferencia.save()

            messages.success(request, 'Transferência criada!')
            return redirect('ERP_ServicesBI:transferencia_list')
    else:
        form = TransferenciaEstoqueForm()

    context = {
        'form': form,
        'depositos': Deposito.objects.filter(ativo=True),
        'produtos': Produto.objects.filter(ativo=True),
    }
    return render(request, 'estoque/transferencia_form.html', context)


@login_required
def produtos_estoque_baixo(request):
    """Lista de produtos com estoque baixo."""
    produtos = Produto.objects.filter(
        estoque_atual__lte=F('estoque_minimo'),
        ativo=True
    ).select_related('categoria').order_by('estoque_atual')

    paginator = Paginator(produtos, 25)
    page_obj = paginator.get_page(request.GET.get('page'))

    context = {
        'produtos': page_obj,
        'page_obj': page_obj,
        'total': produtos.count(),
    }
    return render(request, 'estoque/produtos_estoque_baixo.html', context)


@login_required
def relatorio_estoque(request):
    """Relatório de saldo de estoque por depósito."""
    deposito_id = request.GET.get('deposito')
    produto_id = request.GET.get('produto')

    saldos = []
    produto_selecionado = None
    deposito_selecionado = None

    if produto_id:
        try:
            produto_selecionado = Produto.objects.get(pk=produto_id)
            saldos_query = SaldoEstoque.objects.filter(produto=produto_selecionado)

            if deposito_id:
                saldos_query = saldos_query.filter(deposito_id=deposito_id)
                deposito_selecionado = Deposito.objects.filter(pk=deposito_id).first()

            saldos = saldos_query.select_related('deposito', 'produto')
        except Produto.DoesNotExist:
            messages.error(request, 'Produto não encontrado.')

    if not produto_id and not deposito_id:
        saldos = SaldoEstoque.objects.filter(
            quantidade__lte=F('produto__estoque_minimo')
        ).select_related('deposito', 'produto')[:100]

    context = {
        'saldos': saldos,
        'produtos': Produto.objects.filter(ativo=True).order_by('descricao'),
        'depositos': Deposito.objects.filter(ativo=True).order_by('nome'),
        'produto_selecionado': produto_selecionado,
        'deposito_selecionado': deposito_selecionado,
    }
    return render(request, 'estoque/relatorio_estoque.html', context)


@login_required
def consulta_saldo(request):
    """View para consulta rápida de saldo de estoque."""
    produto_id = request.GET.get('produto')
    deposito_id = request.GET.get('deposito')

    saldos = []
    produto_selecionado = None
    deposito_selecionado = None

    if produto_id:
        try:
            produto_selecionado = Produto.objects.get(pk=produto_id)
            saldos_query = SaldoEstoque.objects.filter(produto=produto_selecionado)

            if deposito_id:
                saldos_query = saldos_query.filter(deposito_id=deposito_id)
                deposito_selecionado = Deposito.objects.filter(pk=deposito_id).first()

            saldos = saldos_query.select_related('deposito', 'produto')
        except Produto.DoesNotExist:
            messages.error(request, 'Produto não encontrado.')

    context = {
        'saldos': saldos,
        'produtos': Produto.objects.filter(ativo=True).order_by('descricao'),
        'depositos': Deposito.objects.filter(ativo=True).order_by('nome'),
        'produto_selecionado': produto_selecionado,
        'deposito_selecionado': deposito_selecionado,
    }
    return render(request, 'estoque/consulta_saldo.html', context)


@login_required
@require_GET
def api_estoque_saldo(request):
    """API para consultar saldo de estoque via query parameters."""
    produto_id = request.GET.get('produto_id')
    deposito_id = request.GET.get('deposito_id')

    if not produto_id:
        return JsonResponse({'success': False, 'message': 'produto_id obrigatório'}, status=400)

    try:
        produto = Produto.objects.get(pk=produto_id)
        saldo_query = SaldoEstoque.objects.filter(produto=produto)

        if deposito_id:
            saldo_query = saldo_query.filter(deposito_id=deposito_id)
            saldo = saldo_query.first()
            return JsonResponse({
                'success': True,
                'produto_id': produto_id,
                'deposito_id': deposito_id,
                'saldo': float(saldo.quantidade) if saldo else 0,
                'preco_custo': float(saldo.produto.preco_custo or 0) if saldo else float(produto.preco_custo or 0),
            })
        else:
            saldos = saldo_query.select_related('deposito')
            return JsonResponse({
                'success': True,
                'produto_id': produto_id,
                'saldo_total': float(sum(s.quantidade for s in saldos)),
                'depositos': [
                    {
                        'deposito_id': s.deposito_id,
                        'deposito_nome': s.deposito.nome,
                        'saldo': float(s.quantidade),
                        'preco_custo': float(s.produto.preco_custo or 0),
                    }
                    for s in saldos
                ]
            })
    except Produto.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Produto não encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erro: {str(e)}'}, status=500)


@login_required
@require_GET
def api_produtos_busca(request):
    """API para buscar produtos por descrição/código."""
    termo = request.GET.get('termo', '').strip()[:100]
    limite = int(request.GET.get('limite', 20))

    if not termo or len(termo) < 2:
        return JsonResponse({'success': False, 'message': 'Termo muito curto'}, status=400)

    try:
        produtos = Produto.objects.filter(
            Q(descricao__icontains=termo) | Q(codigo__icontains=termo),
            ativo=True
        ).values('id', 'codigo', 'descricao', 'preco_venda', 'estoque_minimo')[:limite]

        return JsonResponse({
            'success': True,
            'produtos': list(produtos)
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erro: {str(e)}'}, status=500)


@login_required
@require_POST
def deposito_create_ajax(request):
    """Cria um novo depósito via AJAX para modal."""
    if not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': False, 'message': 'Requisição inválida'}, status=400)

    nome = request.POST.get('nome', '').strip()
    descricao = request.POST.get('descricao', '').strip()

    if not nome:
        return JsonResponse({'success': False, 'message': 'Nome do depósito é obrigatório'})

    if Deposito.objects.filter(nome__iexact=nome).exists():
        return JsonResponse({'success': False, 'message': 'Já existe um depósito com este nome'})

    try:
        deposito = Deposito.objects.create(nome=nome, descricao=descricao, ativo=True)
        return JsonResponse({
            'success': True,
            'id': deposito.id,
            'nome': deposito.nome,
            'message': 'Depósito criado com sucesso!'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erro ao criar depósito: {str(e)}'})


@login_required
def relatorio_movimentacao(request):
    """Relatório de movimentações de estoque por período."""
    from datetime import date, timedelta
    data_inicio = _parse_date(request.GET.get('data_inicio')) or (date.today() - timedelta(days=30))
    data_fim = _parse_date(request.GET.get('data_fim')) or date.today()
    produto_id = request.GET.get('produto')
    tipo_movimento = request.GET.get('tipo')

    movimentacoes = MovimentacaoEstoque.objects.filter(
        data__range=[data_inicio, data_fim]
    ).select_related('produto', 'deposito_origem', 'deposito_destino')

    if produto_id:
        movimentacoes = movimentacoes.filter(produto_id=produto_id)

    if tipo_movimento:
        movimentacoes = movimentacoes.filter(tipo=tipo_movimento)

    movimentacoes = movimentacoes.order_by('-data')

    context = {
        'movimentacoes': movimentacoes,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'total': movimentacoes.count(),
        'produtos': Produto.objects.filter(ativo=True).order_by('descricao'),
        'produto_id': produto_id,
        'tipo_movimento': tipo_movimento,
    }
    return render(request, 'estoque/relatorio_movimentacao.html', context)


# ============================================================================
# FIM DO ARQUIVO COMPLEMENTAR
# ============================================================================
# =============================================================================
# ALIASES PARA COMPATIBILIDADE COM URLS
# =============================================================================
# Algumas URLs usam nomes diferentes das funções

vendedor_excluir = vendedor_delete